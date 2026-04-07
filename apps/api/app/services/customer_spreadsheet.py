from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class CustomerSheetColumn:
    key: str
    header: str
    required_for_new: bool = False
    note: str = ""
    width: int = 18


CUSTOMER_SHEET_COLUMNS: list[CustomerSheetColumn] = [
    CustomerSheetColumn("customer_id", "客户ID", note="导出后回填更新时优先按它匹配；新增可留空", width=10),
    CustomerSheetColumn("customer_code", "客户编号", note="可用于更新匹配；新增可留空自动生成", width=14),
    CustomerSheetColumn("name", "公司名称", required_for_new=True, note="新增必填", width=24),
    CustomerSheetColumn("contact_name", "联系人", note="留空时默认同公司名称", width=16),
    CustomerSheetColumn("phone", "联系电话", note="可选填", width=16),
    CustomerSheetColumn("status", "客户状态", note="默认 ACTIVE", width=12),
    CustomerSheetColumn("accountant_username", "会计账号", required_for_new=True, note="填写系统里的会计用户名；新增必填", width=14),
    CustomerSheetColumn("grade", "等级", note="例如 A / B / C", width=10),
    CustomerSheetColumn("region", "地区", width=14),
    CustomerSheetColumn("country", "国家/地区", width=14),
    CustomerSheetColumn("service_start_text", "服务开始", note="支持 2025.01 或 2025-01-01", width=14),
    CustomerSheetColumn("company_nature", "公司性质", width=14),
    CustomerSheetColumn("service_mode", "服务方式", width=14),
    CustomerSheetColumn("contact_wechat", "微信", width=14),
    CustomerSheetColumn("other_contact", "其他联系方式", width=20),
    CustomerSheetColumn("main_business", "服务项目/主营", note="例如 代账、海外注册、咨询", width=28),
    CustomerSheetColumn("source", "来源", note="留空时默认 Excel导入", width=16),
    CustomerSheetColumn("intro", "介绍人", width=16),
    CustomerSheetColumn("fee_standard", "收费标准", width=18),
    CustomerSheetColumn("first_billing_period", "首次收费期间", width=16),
    CustomerSheetColumn("reminder_value", "提醒值", width=14),
    CustomerSheetColumn("next_reminder_at", "下次提醒日期", note="格式 YYYY-MM-DD，可留空", width=16),
    CustomerSheetColumn("notes", "备注", width=28),
]


HEADER_TO_KEY = {item.header: item.key for item in CUSTOMER_SHEET_COLUMNS}
KEY_TO_HEADER = {item.key: item.header for item in CUSTOMER_SHEET_COLUMNS}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="EAF2FF")
INFO_FILL = PatternFill(fill_type="solid", fgColor="FFF4D6")


def _format_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _apply_sheet_style(ws) -> None:
    ws.freeze_panes = "A2"
    for index, column in enumerate(CUSTOMER_SHEET_COLUMNS, start=1):
        cell = ws.cell(row=1, column=index)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(index)].width = column.width
    ws.row_dimensions[1].height = 24


def build_customer_template_bytes() -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "客户导入模板"
    ws.append([item.header for item in CUSTOMER_SHEET_COLUMNS])
    _apply_sheet_style(ws)

    guide = workbook.create_sheet("填写说明")
    guide.append(["字段", "是否新增必填", "说明"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = INFO_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 14
    guide.column_dimensions["C"].width = 58
    for item in CUSTOMER_SHEET_COLUMNS:
        guide.append([
            item.header,
            "是" if item.required_for_new else "否",
            item.note or "可选填",
        ])
    guide.append([])
    guide.append(["导入规则", "", "有客户ID优先按客户ID更新；否则按客户编号更新；两者都没有则按新增处理。"])
    guide.append(["覆盖规则", "", "更新已有客户时，空白单元格默认不覆盖已有数据。"])
    guide.append(["权限规则", "", "会计账号必须填写系统用户名；经理只能导入给自己或直属会计。"])
    guide.append(["编号规则", "", "新增客户若不填客户编号，会自动生成连续编号。"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_customer_export_bytes(rows: Iterable[dict[str, object]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "客户列表导出"
    ws.append([item.header for item in CUSTOMER_SHEET_COLUMNS])
    _apply_sheet_style(ws)
    for row in rows:
        ws.append([_format_cell_value(row.get(item.key, "")) for item in CUSTOMER_SHEET_COLUMNS])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_customer_import_file(filename: str, content: bytes) -> list[tuple[int, dict[str, str]]]:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return _parse_customer_csv_rows(content)
    return _parse_customer_xlsx_rows(content)


def _parse_customer_xlsx_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    ws = workbook.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    return _rows_from_values(rows)


def _parse_customer_csv_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = [tuple(item for item in line) for line in reader]
    if not rows:
        return []
    return _rows_from_values(rows)


def _rows_from_values(rows: list[tuple[object, ...] | list[object]]) -> list[tuple[int, dict[str, str]]]:
    headers = [_format_cell_value(item) for item in rows[0]]
    missing = [item.header for item in CUSTOMER_SHEET_COLUMNS if item.header not in headers]
    if missing:
        raise ValueError(f"缺少模板列：{'、'.join(missing)}")

    header_indexes = {HEADER_TO_KEY[header]: idx for idx, header in enumerate(headers) if header in HEADER_TO_KEY}
    parsed_rows: list[tuple[int, dict[str, str]]] = []
    for excel_row_number, raw_row in enumerate(rows[1:], start=2):
        payload = {
            key: _format_cell_value(raw_row[index]) if index < len(raw_row) else ""
            for key, index in header_indexes.items()
        }
        if not any(value for value in payload.values()):
            continue
        parsed_rows.append((excel_row_number, payload))
    return parsed_rows
