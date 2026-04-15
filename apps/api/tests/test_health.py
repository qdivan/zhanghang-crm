import os
from datetime import datetime
from datetime import date, timedelta
from io import BytesIO
from types import SimpleNamespace
from urllib.parse import parse_qs, urlparse
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.api.v1 import auth as auth_api
from app.core.config import settings
from app.core.security import create_access_token
from app.main import app
from app.db.session import SessionLocal
from app.models import BillingRecord, SsoBindingConflict, User
from app.services import sso as sso_service
from app.services.customer_spreadsheet import CUSTOMER_SHEET_COLUMNS

DEMO_PASSWORD = os.environ.get("BOOTSTRAP_DEMO_PASSWORD", "Daizhang#2026!")


def build_customer_import_bytes(rows: list[dict[str, str]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "客户导入模板"
    ws.append([item.header for item in CUSTOMER_SHEET_COLUMNS])
    for row in rows:
        ws.append([row.get(item.key, "") for item in CUSTOMER_SHEET_COLUMNS])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_health():
    with TestClient(app) as client:
        response = client.get("/api/v1/health")
        assert response.status_code == 200
        assert response.json() == {"status": "ok"}


def test_department_manager_can_manage_only_direct_subordinate_billing():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"include_inactive": True})
        assert users_resp.status_code == 200
        users = users_resp.json()
        manager = next(item for item in users if item["username"] == "manager")
        accountant = next(item for item in users if item["username"] == "accountant")
        accountant2 = next(item for item in users if item["username"] == "accountant2")

        bind_subordinate = client.patch(
            f"/api/v1/users/{accountant2['id']}",
            headers=owner_headers,
            json={"manager_user_id": manager["id"]},
        )
        assert bind_subordinate.status_code == 200

        clear_non_subordinate = client.patch(
            f"/api/v1/users/{accountant['id']}",
            headers=owner_headers,
            json={"manager_user_id": None},
        )
        assert clear_non_subordinate.status_code == 200

        subordinate_lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "经理下属收费客户-A",
                "grade": "待下单",
                "contact_name": "下属联系人",
                "phone": "13910001001",
                "region": "青岛",
            },
        )
        assert subordinate_lead_resp.status_code == 201
        subordinate_lead_id = subordinate_lead_resp.json()["id"]

        subordinate_convert_resp = client.post(
            f"/api/v1/leads/{subordinate_lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant2["id"]},
        )
        assert subordinate_convert_resp.status_code == 200
        subordinate_customer_id = subordinate_convert_resp.json()["customer"]["id"]

        non_subordinate_lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "经理无权收费客户-B",
                "grade": "待下单",
                "contact_name": "非下属联系人",
                "phone": "13910001002",
                "region": "上海",
            },
        )
        assert non_subordinate_lead_resp.status_code == 201
        non_subordinate_lead_id = non_subordinate_lead_resp.json()["id"]

        non_subordinate_convert_resp = client.post(
            f"/api/v1/leads/{non_subordinate_lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert non_subordinate_convert_resp.status_code == 200
        non_subordinate_customer_id = non_subordinate_convert_resp.json()["customer"]["id"]

        create_subordinate_billing = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": subordinate_customer_id,
                "charge_category": "代账",
                "charge_mode": "PERIODIC",
                "amount_basis": "YEARLY",
                "summary": "部门经理可管理的收费单",
                "total_fee": 3600,
                "monthly_fee": 300,
                "collection_start_date": "2026-03-01",
                "due_month": "2027-02-28",
                "payment_method": "后收",
                "note": "直属下属客户",
            },
        )
        assert create_subordinate_billing.status_code == 201

        create_non_subordinate_billing = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": non_subordinate_customer_id,
                "charge_category": "注册",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "部门经理无权管理的收费单",
                "total_fee": 800,
                "monthly_fee": 0,
                "collection_start_date": "2026-03-05",
                "due_month": "2026-03-05",
                "payment_method": "后收",
                "note": "非直属下属客户",
            },
        )
        assert create_non_subordinate_billing.status_code == 201

        manager_login = client.post(
            "/api/v1/auth/login",
            json={"username": "manager", "password": DEMO_PASSWORD},
        )
        assert manager_login.status_code == 200
        manager_headers = {"Authorization": f"Bearer {manager_login.json()['access_token']}"}

        subordinate_users_resp = client.get(
            "/api/v1/users",
            headers=manager_headers,
            params={"role": "ACCOUNTANT", "include_inactive": True},
        )
        assert subordinate_users_resp.status_code == 200
        subordinate_usernames = {item["username"] for item in subordinate_users_resp.json()}
        assert "accountant2" in subordinate_usernames
        assert "accountant" not in subordinate_usernames

        customer_list_resp = client.get("/api/v1/customers", headers=manager_headers)
        assert customer_list_resp.status_code == 200
        customer_names = {item["name"] for item in customer_list_resp.json()}
        assert "经理下属收费客户-A" in customer_names
        assert "经理无权收费客户-B" not in customer_names

        billing_list_resp = client.get("/api/v1/billing-records", headers=manager_headers)
        assert billing_list_resp.status_code == 200
        billing_customer_names = {item["customer_name"] for item in billing_list_resp.json()}
        assert "经理下属收费客户-A" in billing_customer_names
        assert "经理无权收费客户-B" not in billing_customer_names

        manager_create_subordinate_billing = client.post(
            "/api/v1/billing-records",
            headers=manager_headers,
            json={
                "customer_id": subordinate_customer_id,
                "charge_category": "咨询",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "经理代下属补录收费单",
                "total_fee": 600,
                "monthly_fee": 0,
                "collection_start_date": "2026-03-10",
                "due_month": "2026-03-10",
                "payment_method": "后收",
            },
        )
        assert manager_create_subordinate_billing.status_code == 201

        manager_create_non_subordinate_billing = client.post(
            "/api/v1/billing-records",
            headers=manager_headers,
            json={
                "customer_id": non_subordinate_customer_id,
                "charge_category": "咨询",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "经理越权收费单",
                "total_fee": 600,
                "monthly_fee": 0,
                "collection_start_date": "2026-03-10",
                "due_month": "2026-03-10",
                "payment_method": "后收",
            },
        )
        assert manager_create_non_subordinate_billing.status_code == 403


def test_customer_timeline_open_event_can_remind_and_complete():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"include_inactive": True})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "CONVERSION",
                "name": "客户事项提醒测试-A",
                "grade": "已签合同/待交费",
                "contact_name": "王小姐",
                "phone": "13910002001",
                "region": "香港",
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        reminder_date = (date.today() + timedelta(days=2)).isoformat()
        create_event_resp = client.post(
            f"/api/v1/customers/{customer_id}/timeline-events",
            headers=headers,
            json={
                "occurred_at": date.today().isoformat(),
                "event_type": "DELIVERY",
                "status": "OPEN",
                "reminder_at": reminder_date,
                "content": "等待客户补交香港公司年审资料",
                "note": "先催一次",
            },
        )
        assert create_event_resp.status_code == 201
        event_id = create_event_resp.json()["id"]
        assert create_event_resp.json()["status"] == "OPEN"

        system_todos_resp = client.get("/api/v1/dashboard/system-todos", headers=headers, params={"limit": 100})
        assert system_todos_resp.status_code == 200
        matched = [
            item for item in system_todos_resp.json()
            if item["module"] == "CUSTOMER" and item["action_path"] == f"/customers/{customer_id}"
        ]
        assert matched

        complete_resp = client.patch(
            f"/api/v1/customers/{customer_id}/timeline-events/{event_id}",
            headers=headers,
            json={
                "status": "DONE",
                "completed_at": date.today().isoformat(),
                "result": "资料已收到并已提交办理",
            },
        )
        assert complete_resp.status_code == 200
        assert complete_resp.json()["status"] == "DONE"
        assert complete_resp.json()["result"] == "资料已收到并已提交办理"

        detail_resp = client.get(f"/api/v1/customers/{customer_id}", headers=headers)
        assert detail_resp.status_code == 200
        customer_events = [
            item for item in detail_resp.json()["timeline"]
            if item["source_type"] == "CUSTOMER_EVENT" and item["source_id"] == event_id
        ]
        assert customer_events
        assert customer_events[0]["status"] == "DONE"
        assert customer_events[0]["result"] == "资料已收到并已提交办理"


def test_apply_hk_company_template_creates_customer_followup_tasks():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"include_inactive": True})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "CONVERSION",
                "name": "香港模板客户-B",
                "grade": "已签合同/待交费",
                "contact_name": "林先生",
                "phone": "13910002002",
                "region": "香港",
                "service_start_text": date.today().isoformat(),
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        apply_resp = client.post(
            f"/api/v1/customers/{customer_id}/timeline-templates/hk-company",
            headers=headers,
        )
        assert apply_resp.status_code == 201
        items = apply_resp.json()
        assert len(items) == 3
        assert {item["status"] for item in items} == {"OPEN"}
        assert {item["template_key"] for item in items} == {"HK_COMPANY"}

        duplicate_resp = client.post(
            f"/api/v1/customers/{customer_id}/timeline-templates/hk-company",
            headers=headers,
        )
        assert duplicate_resp.status_code == 400


def test_common_library_supports_internal_and_public_visibility():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        create_resp = client.post(
            "/api/v1/common-library-items",
            headers=headers,
            json={
                "module_type": "EXTENSION_A",
                "visibility": "PUBLIC",
                "category": "官网知识库",
                "title": "退税资料清单",
                "content": "公开版退税资料清单示例",
            },
        )
        assert create_resp.status_code == 201
        item_id = create_resp.json()["id"]
        assert create_resp.json()["visibility"] == "PUBLIC"

        public_resp = client.get("/api/v1/common-library-items/public")
        assert public_resp.status_code == 200
        public_ids = {item["id"] for item in public_resp.json()}
        assert item_id in public_ids

        internal_list_resp = client.get(
            "/api/v1/common-library-items",
            headers=headers,
            params={"visibility": "INTERNAL"},
        )
        assert internal_list_resp.status_code == 200
        internal_ids = {item["id"] for item in internal_list_resp.json()}
        assert item_id not in internal_ids


def test_login_and_me():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        token = login_response.json()["access_token"]

        me_response = client.get(
            "/api/v1/auth/me",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert me_response.status_code == 200
        assert me_response.json()["username"] == "boss"


def test_security_settings_default_and_update():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        get_resp = client.get("/api/v1/admin/security-settings", headers=headers)
        assert get_resp.status_code == 200
        assert get_resp.json()["local_ip_lock_enabled"] is True
        assert get_resp.json()["local_ip_lock_window_minutes"] == 5
        assert get_resp.json()["local_ip_lock_max_attempts"] == 20

        update_resp = client.put(
            "/api/v1/admin/security-settings",
            headers=headers,
            json={
                "local_ip_lock_enabled": True,
                "local_ip_lock_window_minutes": 7,
                "local_ip_lock_max_attempts": 9,
            },
        )
        assert update_resp.status_code == 200
        updated = update_resp.json()
        assert updated["local_ip_lock_window_minutes"] == 7
        assert updated["local_ip_lock_max_attempts"] == 9


def test_local_login_ip_lock_blocks_after_threshold_and_success_clears_counter():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        settings_resp = client.put(
            "/api/v1/admin/security-settings",
            headers=headers,
            json={
                "local_ip_lock_enabled": True,
                "local_ip_lock_window_minutes": 5,
                "local_ip_lock_max_attempts": 3,
            },
        )
        assert settings_resp.status_code == 200

        blocked_ip_headers = {"x-forwarded-for": "198.51.100.20"}
        for _ in range(2):
            fail_resp = client.post(
                "/api/v1/auth/login",
                headers=blocked_ip_headers,
                json={"username": "boss", "password": "wrong-password"},
            )
            assert fail_resp.status_code == 401

        threshold_resp = client.post(
            "/api/v1/auth/login",
            headers=blocked_ip_headers,
            json={"username": "boss", "password": "wrong-password"},
        )
        assert threshold_resp.status_code == 429
        assert "已锁定 5 分钟" in threshold_resp.json()["detail"]

        blocked_valid_resp = client.post(
            "/api/v1/auth/login",
            headers=blocked_ip_headers,
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert blocked_valid_resp.status_code == 429

        safe_ip_headers = {"x-forwarded-for": "198.51.100.30"}
        first_fail_resp = client.post(
            "/api/v1/auth/login",
            headers=safe_ip_headers,
            json={"username": "boss", "password": "wrong-password"},
        )
        assert first_fail_resp.status_code == 401

        success_resp = client.post(
            "/api/v1/auth/login",
            headers=safe_ip_headers,
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert success_resp.status_code == 200

        after_success_fail_resp = client.post(
            "/api/v1/auth/login",
            headers=safe_ip_headers,
            json={"username": "boss", "password": "wrong-password"},
        )
        assert after_success_fail_resp.status_code == 401


def test_billing_records():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        list_response = client.get("/api/v1/billing-records", headers=headers)
        assert list_response.status_code == 200
        assert len(list_response.json()) >= 1

        customers_response = client.get("/api/v1/customers", headers=headers)
        assert customers_response.status_code == 200
        assert len(customers_response.json()) >= 1
        customer_id = customers_response.json()[0]["id"]

        create_response = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "serial_no": 99,
                "customer_id": customer_id,
                "total_fee": 1000,
                "monthly_fee": 100,
                "billing_cycle_text": "2026/2/26收（2026.1-2026.12）",
                "collection_start_date": "2026-01-01",
                "due_month": "2026-12-31",
                "payment_method": "预收",
                "note": "测试记录",
            },
        )
        assert create_response.status_code == 201
        assert isinstance(create_response.json()["serial_no"], int)
        assert create_response.json()["collection_start_date"] == "2026-01-01"


def test_billing_records_support_contact_filter_and_keyword():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        unique_contact_name = "联系人筛选_唯一_20260228"
        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "联系人筛选测试客户",
                "contact_name": unique_contact_name,
                "phone": "13800138033",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "total_fee": 1888,
                "monthly_fee": 188,
                "billing_cycle_text": "联系人筛选账期",
                "collection_start_date": "2026-01-15",
                "due_month": "2026-12-20",
                "payment_method": "后收",
                "note": "联系人筛选备注",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        by_contact_resp = client.get(
            "/api/v1/billing-records",
            headers=headers,
            params={"contact_name": unique_contact_name},
        )
        assert by_contact_resp.status_code == 200
        by_contact_records = by_contact_resp.json()
        assert any(item["id"] == record_id for item in by_contact_records)
        assert all("customer_contact_name" in item for item in by_contact_records)

        by_keyword_resp = client.get(
            "/api/v1/billing-records",
            headers=headers,
            params={"keyword": unique_contact_name},
        )
        assert by_keyword_resp.status_code == 200
        by_keyword_records = by_keyword_resp.json()
        target = next(item for item in by_keyword_records if item["id"] == record_id)
        assert target["customer_contact_name"] == unique_contact_name
        assert target["collection_start_date"] == "2026-01-15"


def test_billing_defaults_for_periodic_and_one_time_modes():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        customers_response = client.get("/api/v1/customers", headers=headers)
        assert customers_response.status_code == 200
        customer_id = customers_response.json()[0]["id"]

        periodic_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "amount_basis": "MONTHLY",
                "period_start_month": "2026-03",
                "total_fee": 1200,
                "payment_method": "后收",
            },
        )
        assert periodic_resp.status_code == 201
        assert periodic_resp.json()["period_start_month"] == "2026-03"
        assert periodic_resp.json()["period_end_month"] == "2027-02"
        assert periodic_resp.json()["collection_start_date"] == "2026-03-01"
        assert periodic_resp.json()["due_month"] == "2027-02-28"

        one_time_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "ONE_TIME",
                "total_fee": 800,
                "payment_method": "后收",
            },
        )
        assert one_time_resp.status_code == 201
        assert one_time_resp.json()["amount_basis"] == "ONE_TIME"
        assert one_time_resp.json()["due_month"] == date.today().isoformat()


def test_billing_batch_create_supports_multiple_rows_for_same_customer():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        customers_response = client.get("/api/v1/customers", headers=headers)
        assert customers_response.status_code == 200
        customer_id = customers_response.json()[0]["id"]

        batch_resp = client.post(
            "/api/v1/billing-records/batch",
            headers=headers,
            json={
                "records": [
                    {
                        "customer_id": customer_id,
                        "charge_category": "代账",
                        "charge_mode": "PERIODIC",
                        "amount_basis": "MONTHLY",
                        "summary": "批量代账服务",
                        "total_fee": 3600,
                        "monthly_fee": 300,
                        "period_start_month": "2026-04",
                        "payment_method": "后收",
                    },
                    {
                        "customer_id": customer_id,
                        "charge_category": "注册",
                        "charge_mode": "ONE_TIME",
                        "summary": "批量注册服务",
                        "total_fee": 1800,
                        "payment_method": "预收",
                    },
                ]
            },
        )
        assert batch_resp.status_code == 201
        records = batch_resp.json()
        assert len(records) == 2
        assert records[0]["customer_id"] == customer_id
        assert records[1]["customer_id"] == customer_id
        assert records[0]["period_start_month"] == "2026-04"
        assert records[0]["period_end_month"] == "2027-03"
        assert records[0]["collection_start_date"] == "2026-04-01"
        assert records[0]["due_month"] == "2027-03-31"
        assert records[1]["due_month"] == date.today().isoformat()
        assert records[0]["serial_no"] + 1 == records[1]["serial_no"]


def test_accountant_can_only_access_own_billing_records():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        users = users_resp.json()
        accountant = next(item for item in users if item["username"] == "accountant")
        accountant2 = next(item for item in users if item["username"] == "accountant2")

        lead_owner_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "收费可见性-会计1",
                "contact_name": "张一",
                "phone": "13800138031",
            },
        )
        assert lead_owner_resp.status_code == 201
        lead_owner_id = lead_owner_resp.json()["id"]

        convert_owner = client.post(
            f"/api/v1/leads/{lead_owner_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_owner.status_code == 200
        customer_owner_id = convert_owner.json()["customer"]["id"]

        lead_other_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "收费可见性-会计2",
                "contact_name": "张二",
                "phone": "13800138032",
            },
        )
        assert lead_other_resp.status_code == 201
        lead_other_id = lead_other_resp.json()["id"]

        convert_other = client.post(
            f"/api/v1/leads/{lead_other_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant2["id"]},
        )
        assert convert_other.status_code == 200
        customer_other_id = convert_other.json()["customer"]["id"]

        record_owner_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1201,
                "customer_id": customer_owner_id,
                "total_fee": 1000,
                "monthly_fee": 100,
                "billing_cycle_text": "会计1账单",
                "due_month": "2026-12-31",
                "payment_method": "后收",
            },
        )
        assert record_owner_resp.status_code == 201
        record_owner_id = record_owner_resp.json()["id"]

        record_other_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1202,
                "customer_id": customer_other_id,
                "total_fee": 2000,
                "monthly_fee": 200,
                "billing_cycle_text": "会计2账单",
                "due_month": "2026-12-31",
                "payment_method": "后收",
            },
        )
        assert record_other_resp.status_code == 201
        record_other_id = record_other_resp.json()["id"]

        accountant2_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant2", "password": DEMO_PASSWORD},
        )
        assert accountant2_login.status_code == 200
        accountant2_headers = {"Authorization": f"Bearer {accountant2_login.json()['access_token']}"}

        visible_records_resp = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert visible_records_resp.status_code == 200
        visible_record_ids = {item["id"] for item in visible_records_resp.json()}
        assert record_other_id in visible_record_ids
        assert record_owner_id not in visible_record_ids

        summary_resp = client.get("/api/v1/billing-records/summary", headers=accountant2_headers)
        assert summary_resp.status_code == 200
        assert summary_resp.json()["total_records"] == 1
        assert summary_resp.json()["total_fee"] == 2000

        forbidden_activity_list = client.get(
            f"/api/v1/billing-records/{record_owner_id}/activities",
            headers=accountant2_headers,
        )
        assert forbidden_activity_list.status_code == 403

        forbidden_activity_create = client.post(
            f"/api/v1/billing-records/{record_owner_id}/activities",
            headers=accountant2_headers,
            json={
                "activity_type": "REMINDER",
                "occurred_at": "2026-02-26",
                "amount": 0,
                "content": "无权限客户不应可催收",
            },
        )
        assert forbidden_activity_create.status_code == 403


def test_billing_assignment_grants_read_visibility_to_executor():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        users = users_resp.json()
        accountant = next(item for item in users if item["username"] == "accountant")
        accountant2 = next(item for item in users if item["username"] == "accountant2")

        lead_owner_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "执行分派可见性客户A",
                "contact_name": "执行A",
                "phone": "13800138051",
            },
        )
        assert lead_owner_resp.status_code == 201
        lead_owner_id = lead_owner_resp.json()["id"]

        convert_owner = client.post(
            f"/api/v1/leads/{lead_owner_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_owner.status_code == 200
        customer_owner_id = convert_owner.json()["customer"]["id"]

        record_owner_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1401,
                "customer_id": customer_owner_id,
                "total_fee": 5000,
                "monthly_fee": 500,
                "billing_cycle_text": "执行分派账单",
                "due_month": "2026-12-31",
                "payment_method": "后收",
            },
        )
        assert record_owner_resp.status_code == 201
        record_owner_id = record_owner_resp.json()["id"]

        accountant2_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant2", "password": DEMO_PASSWORD},
        )
        assert accountant2_login.status_code == 200
        accountant2_headers = {"Authorization": f"Bearer {accountant2_login.json()['access_token']}"}

        before_assign = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert before_assign.status_code == 200
        assert all(item["id"] != record_owner_id for item in before_assign.json())

        create_assignment = client.post(
            f"/api/v1/billing-records/{record_owner_id}/assignees",
            headers=owner_headers,
            json={
                "assignee_user_id": accountant2["id"],
                "assignment_role": "REGISTRATION",
                "note": "负责注册办理，需看费用标准",
            },
        )
        assert create_assignment.status_code == 201
        assignment_id = create_assignment.json()["id"]

        assignment_list = client.get(f"/api/v1/billing-records/{record_owner_id}/assignees", headers=accountant2_headers)
        assert assignment_list.status_code == 200
        assert any(item["id"] == assignment_id for item in assignment_list.json())

        after_assign = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert after_assign.status_code == 200
        assert any(item["id"] == record_owner_id for item in after_assign.json())

        read_activities = client.get(f"/api/v1/billing-records/{record_owner_id}/activities", headers=accountant2_headers)
        assert read_activities.status_code == 200

        write_activity = client.post(
            f"/api/v1/billing-records/{record_owner_id}/activities",
            headers=accountant2_headers,
            json={
                "activity_type": "REMINDER",
                "occurred_at": "2026-02-26",
                "amount": 0,
                "content": "执行人员只读，不应能新增催收",
            },
        )
        assert write_activity.status_code == 403

        create_execution_log = client.post(
            f"/api/v1/billing-records/{record_owner_id}/execution-logs",
            headers=accountant2_headers,
            json={
                "occurred_at": "2026-02-26",
                "progress_type": "MILESTONE",
                "content": "已完成注册资料提交",
                "next_action": "等待工商审批结果",
                "due_date": "2026-03-03",
                "note": "执行分派人员可写执行进度",
            },
        )
        assert create_execution_log.status_code == 201

        list_execution_log = client.get(
            f"/api/v1/billing-records/{record_owner_id}/execution-logs",
            headers=accountant2_headers,
        )
        assert list_execution_log.status_code == 200
        assert any(item["content"] == "已完成注册资料提交" for item in list_execution_log.json())

        disable_assignment = client.patch(
            f"/api/v1/billing-records/{record_owner_id}/assignees/{assignment_id}",
            headers=owner_headers,
            json={"is_active": False},
        )
        assert disable_assignment.status_code == 200
        assert disable_assignment.json()["is_active"] is False

        after_disable = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert after_disable.status_code == 200
        assert all(item["id"] != record_owner_id for item in after_disable.json())

        write_execution_after_disable = client.post(
            f"/api/v1/billing-records/{record_owner_id}/execution-logs",
            headers=accountant2_headers,
            json={
                "occurred_at": "2026-02-27",
                "progress_type": "UPDATE",
                "content": "停用后不应再写入",
            },
        )
        assert write_execution_after_disable.status_code == 403


def test_accountant_read_only_data_grant_for_customer_and_billing():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        users = users_resp.json()
        accountant = next(item for item in users if item["username"] == "accountant")
        accountant2 = next(item for item in users if item["username"] == "accountant2")

        lead_owner_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "临时授权客户A",
                "contact_name": "授权A",
                "phone": "13800138041",
            },
        )
        assert lead_owner_resp.status_code == 201
        lead_owner_id = lead_owner_resp.json()["id"]
        convert_owner = client.post(
            f"/api/v1/leads/{lead_owner_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_owner.status_code == 200
        customer_owner_id = convert_owner.json()["customer"]["id"]

        lead_other_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "临时授权客户B",
                "contact_name": "授权B",
                "phone": "13800138042",
            },
        )
        assert lead_other_resp.status_code == 201
        lead_other_id = lead_other_resp.json()["id"]
        convert_other = client.post(
            f"/api/v1/leads/{lead_other_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant2["id"]},
        )
        assert convert_other.status_code == 200
        customer_other_id = convert_other.json()["customer"]["id"]

        record_owner_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1301,
                "customer_id": customer_owner_id,
                "total_fee": 3000,
                "monthly_fee": 300,
                "billing_cycle_text": "授权A账单",
                "due_month": "2026-12-31",
                "payment_method": "后收",
            },
        )
        assert record_owner_resp.status_code == 201
        record_owner_id = record_owner_resp.json()["id"]

        record_other_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1302,
                "customer_id": customer_other_id,
                "total_fee": 2000,
                "monthly_fee": 200,
                "billing_cycle_text": "授权B账单",
                "due_month": "2026-12-31",
                "payment_method": "后收",
            },
        )
        assert record_other_resp.status_code == 201
        record_other_id = record_other_resp.json()["id"]

        accountant2_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant2", "password": DEMO_PASSWORD},
        )
        assert accountant2_login.status_code == 200
        accountant2_headers = {"Authorization": f"Bearer {accountant2_login.json()['access_token']}"}

        customers_before = client.get("/api/v1/customers", headers=accountant2_headers)
        assert customers_before.status_code == 200
        customer_ids_before = {item["id"] for item in customers_before.json()}
        assert customer_other_id in customer_ids_before
        assert customer_owner_id not in customer_ids_before

        billing_before = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert billing_before.status_code == 200
        billing_ids_before = {item["id"] for item in billing_before.json()}
        assert record_other_id in billing_ids_before
        assert record_owner_id not in billing_ids_before

        create_customer_grant = client.post(
            "/api/v1/admin/data-access-grants",
            headers=owner_headers,
            json={
                "grantee_user_id": accountant2["id"],
                "module": "CUSTOMER",
                "reason": "月底工资核算临时查看客户",
            },
        )
        assert create_customer_grant.status_code == 201

        create_billing_grant = client.post(
            "/api/v1/admin/data-access-grants",
            headers=owner_headers,
            json={
                "grantee_user_id": accountant2["id"],
                "module": "BILLING",
                "reason": "月底工资核算临时查看收费",
            },
        )
        assert create_billing_grant.status_code == 201
        billing_grant_id = create_billing_grant.json()["id"]

        customers_after = client.get("/api/v1/customers", headers=accountant2_headers)
        assert customers_after.status_code == 200
        customer_ids_after = {item["id"] for item in customers_after.json()}
        assert customer_other_id in customer_ids_after
        assert customer_owner_id in customer_ids_after

        foreign_customer_detail = client.get(f"/api/v1/customers/{customer_owner_id}", headers=accountant2_headers)
        assert foreign_customer_detail.status_code == 200

        foreign_customer_update = client.patch(
            f"/api/v1/customers/{customer_owner_id}",
            headers=accountant2_headers,
            json={"name": "不应允许修改"},
        )
        assert foreign_customer_update.status_code == 403

        billing_after = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert billing_after.status_code == 200
        billing_ids_after = {item["id"] for item in billing_after.json()}
        assert record_other_id in billing_ids_after
        assert record_owner_id in billing_ids_after

        foreign_billing_activities = client.get(
            f"/api/v1/billing-records/{record_owner_id}/activities",
            headers=accountant2_headers,
        )
        assert foreign_billing_activities.status_code == 200

        foreign_billing_write = client.post(
            f"/api/v1/billing-records/{record_owner_id}/activities",
            headers=accountant2_headers,
            json={
                "activity_type": "REMINDER",
                "occurred_at": "2026-02-26",
                "amount": 0,
                "content": "临时授权仅只读，不应允许写入",
            },
        )
        assert foreign_billing_write.status_code == 403

        revoke_billing_grant = client.patch(
            f"/api/v1/admin/data-access-grants/{billing_grant_id}",
            headers=owner_headers,
            json={"is_active": False},
        )
        assert revoke_billing_grant.status_code == 200
        assert revoke_billing_grant.json()["is_active"] is False


def test_lead_detail_and_convert_customer_flow():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        create_lead = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "FOLLOWUP",
                "name": "转化测试公司",
                "contact_name": "李经理",
                "phone": "13800138000",
                "country": "跨境电商",
            },
        )
        assert create_lead.status_code == 201
        lead_id = create_lead.json()["id"]

        lead_detail = client.get(f"/api/v1/leads/{lead_id}", headers=headers)
        assert lead_detail.status_code == 200
        assert lead_detail.json()["name"] == "转化测试公司"

        accountants_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert accountants_resp.status_code == 200
        accountant = accountants_resp.json()[0]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={
                "accountant_id": accountant["id"],
                "customer_name": "转化后客户名",
                "customer_contact_name": "转化后联系人",
                "customer_phone": "13911112222",
            },
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        customers_resp = client.get("/api/v1/customers", headers=headers)
        assert customers_resp.status_code == 200
        assert any(item["id"] == customer_id for item in customers_resp.json())

        customer_detail = client.get(f"/api/v1/customers/{customer_id}", headers=headers)
        assert customer_detail.status_code == 200
        assert customer_detail.json()["lead"]["id"] == lead_id
        assert customer_detail.json()["name"] == "转化后客户名"
        assert customer_detail.json()["contact_name"] == "转化后联系人"
        assert customer_detail.json()["phone"] == "13911112222"

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": accountant["username"], "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}
        accountant_followup = client.post(
            f"/api/v1/leads/{lead_id}/followups",
            headers=accountant_headers,
            json={
                "followup_at": "2026-02-26",
                "feedback": "会计已跟进并记录",
                "notes": "测试",
            },
        )
        assert accountant_followup.status_code == 201

        accountant_leads = client.get("/api/v1/leads", headers=accountant_headers)
        assert accountant_leads.status_code == 200
        assert any(item["id"] == lead_id for item in accountant_leads.json())

        unconvert_resp = client.post(f"/api/v1/leads/{lead_id}/unconvert", headers=headers, json={})
        assert unconvert_resp.status_code == 200
        assert unconvert_resp.json()["status"] in {"NEW", "FOLLOWING"}

        customers_after = client.get("/api/v1/customers", headers=headers).json()
        assert all(item["source_lead_id"] != lead_id for item in customers_after)


def test_redevelop_lead_convert_defaults_to_new_linked_customer():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        accountants_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert accountants_resp.status_code == 200
        accountants = accountants_resp.json()
        assert len(accountants) >= 2
        accountant_1 = accountants[0]
        accountant_2 = accountants[1]

        create_base_lead = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "FOLLOWUP",
                "name": "老客二开基准客户",
                "contact_name": "赵老板",
                "phone": "13800138111",
            },
        )
        assert create_base_lead.status_code == 201
        base_lead_id = create_base_lead.json()["id"]

        convert_base = client.post(
            f"/api/v1/leads/{base_lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant_1["id"]},
        )
        assert convert_base.status_code == 200
        base_customer_id = convert_base.json()["customer"]["id"]

        customer_count_before = len(client.get("/api/v1/customers", headers=headers).json())

        create_redevelop_lead = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "REDEVELOP",
                "related_customer_id": base_customer_id,
                "name": "老客二开基准客户",
                "contact_name": "赵老板",
                "phone": "13800138111",
                "source": "老客户二次开发",
                "notes": "新增股权变更服务",
            },
        )
        assert create_redevelop_lead.status_code == 201
        redevelop_lead = create_redevelop_lead.json()
        redevelop_lead_id = redevelop_lead["id"]
        assert redevelop_lead["related_customer_id"] == base_customer_id
        assert redevelop_lead["customer_id"] == base_customer_id

        convert_redevelop = client.post(
            f"/api/v1/leads/{redevelop_lead_id}/convert",
            headers=headers,
            json={
                "accountant_id": accountant_2["id"],
                "customer_name": "老客二开基准客户-升级服务",
            },
        )
        assert convert_redevelop.status_code == 200
        new_customer = convert_redevelop.json()["customer"]
        assert new_customer["id"] != base_customer_id
        assert new_customer["source_customer_id"] == base_customer_id
        assert new_customer["name"] == "老客二开基准客户-升级服务"

        customers_after = client.get("/api/v1/customers", headers=headers).json()
        assert len(customers_after) == customer_count_before + 1
        original_customer = next(item for item in customers_after if item["id"] == base_customer_id)
        created_customer = next(item for item in customers_after if item["id"] == new_customer["id"])
        assert original_customer["assigned_accountant_id"] == accountant_1["id"]
        assert created_customer["assigned_accountant_id"] == accountant_2["id"]
        assert created_customer["source_customer_id"] == base_customer_id

        redevelop_detail = client.get(f"/api/v1/leads/{redevelop_lead_id}", headers=headers)
        assert redevelop_detail.status_code == 200
        assert redevelop_detail.json()["status"] == "CONVERTED"
        assert redevelop_detail.json()["customer_id"] == new_customer["id"]


def test_address_resources():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        list_response = client.get("/api/v1/address-resources", headers=headers)
        assert list_response.status_code == 200

        create_response = client.post(
            "/api/v1/address-resources",
            headers=headers,
            json={
                "category": "注册地址",
                "contact_info": "微信 test001",
                "served_companies": "测试公司A、测试公司B",
                "description": "支持挂靠",
                "next_action": "下周回访",
                "notes": "测试新增",
            },
        )
        assert create_response.status_code == 201
        resource_id = create_response.json()["id"]
        assert create_response.json()["served_companies"] == "测试公司A、测试公司B"
        assert create_response.json()["served_company_count"] == 2

        patch_response = client.patch(
            f"/api/v1/address-resources/{resource_id}",
            headers=headers,
            json={"served_companies": "测试公司A、测试公司C"},
        )
        assert patch_response.status_code == 200
        assert patch_response.json()["served_companies"] == "测试公司A、测试公司C"
        assert [item["company_name"] for item in patch_response.json()["company_items"]] == ["测试公司A", "测试公司C"]

        patch_with_null = client.patch(
            f"/api/v1/address-resources/{resource_id}",
            headers=headers,
            json={"category": None},
        )
        assert patch_with_null.status_code == 400

        empty_create_response = client.post(
            "/api/v1/address-resources",
            headers=headers,
            json={},
        )
        assert empty_create_response.status_code == 422


def test_common_library_items_crud():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        list_response = client.get(
            "/api/v1/common-library-items",
            headers=headers,
            params={"module_type": "TEMPLATE"},
        )
        assert list_response.status_code == 200

        create_response = client.post(
            "/api/v1/common-library-items",
            headers=headers,
            json={
                "module_type": "DIRECTORY",
                "category": "税局",
                "title": "崂山税局",
                "phone": "0532-55556666",
                "address": "青岛市崂山区示例路 88 号",
                "notes": "测试通讯录",
            },
        )
        assert create_response.status_code == 201
        item_id = create_response.json()["id"]
        assert create_response.json()["module_type"] == "DIRECTORY"

        update_response = client.patch(
            f"/api/v1/common-library-items/{item_id}",
            headers=headers,
            json={"notes": "测试通讯录-已更新"},
        )
        assert update_response.status_code == 200
        assert update_response.json()["notes"] == "测试通讯录-已更新"

        delete_response = client.delete(
            f"/api/v1/common-library-items/{item_id}",
            headers=headers,
            params={"confirm_name": "崂山税局"},
        )
        assert delete_response.status_code == 204

        missing_response = client.patch(
            f"/api/v1/common-library-items/{item_id}",
            headers=headers,
            json={"notes": "不存在"},
        )
        assert missing_response.status_code == 404


def test_billing_activities_update_amounts():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}
        customers_response = client.get("/api/v1/customers", headers=headers)
        assert customers_response.status_code == 200
        customer_id = customers_response.json()[0]["id"]

        create_record = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "serial_no": 100,
                "customer_id": customer_id,
                "total_fee": 1000,
                "monthly_fee": 100,
                "billing_cycle_text": "测试周期",
                "due_month": "2026-12-31",
                "payment_method": "后收",
            },
        )
        assert create_record.status_code == 201
        record_id = create_record.json()["id"]

        activity_response = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": "2026-02-26",
                "amount": 300,
                "payment_nature": "ONE_OFF",
                "is_prepay": False,
                "is_settlement": False,
                "content": "收到部分款项",
                "note": "测试收款",
            },
        )
        assert activity_response.status_code == 201

        reminder_with_amount = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=headers,
            json={
                "activity_type": "REMINDER",
                "occurred_at": "2026-02-26",
                "amount": 200,
                "content": "催收不应填写金额",
            },
        )
        assert reminder_with_amount.status_code == 400

        records = client.get("/api/v1/billing-records", headers=headers).json()
        target = next(item for item in records if item["id"] == record_id)
        assert target["received_amount"] == 300
        assert target["outstanding_amount"] == 700
        assert target["status"] == "PARTIAL"


def test_payment_split_suggestion_and_apply_updates_multiple_records():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "分摊收款测试客户",
                "contact_name": "分摊联系人",
                "phone": "13800138061",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_a = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "serial_no": 1501,
                "customer_id": customer_id,
                "total_fee": 2000,
                "monthly_fee": 200,
                "billing_cycle_text": "项目A",
                "due_month": "2026-03-31",
                "payment_method": "后收",
                "summary": "项目A服务费",
            },
        )
        assert record_a.status_code == 201
        record_a_id = record_a.json()["id"]

        record_b = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "serial_no": 1502,
                "customer_id": customer_id,
                "total_fee": 1800,
                "monthly_fee": 180,
                "billing_cycle_text": "项目B",
                "due_month": "2026-06-30",
                "payment_method": "后收",
                "summary": "项目B服务费",
            },
        )
        assert record_b.status_code == 201
        record_b_id = record_b.json()["id"]

        suggest_resp = client.post(
            "/api/v1/billing-records/payments/suggest",
            headers=headers,
            json={
                "customer_id": customer_id,
                "amount": 2500,
                "strategy": "DUE_DATE_ASC",
            },
        )
        assert suggest_resp.status_code == 200
        suggestion = suggest_resp.json()
        assert suggestion["customer_id"] == customer_id
        assert suggestion["suggested_total"] == 2500
        assert suggestion["remaining_amount"] == 0
        assert len(suggestion["allocations"]) >= 2
        first_allocation = suggestion["allocations"][0]
        second_allocation = suggestion["allocations"][1]
        assert first_allocation["billing_record_id"] == record_a_id
        assert second_allocation["billing_record_id"] == record_b_id

        apply_resp = client.post(
            "/api/v1/billing-records/payments",
            headers=headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-02-28",
                "amount": 2500,
                "strategy": "DUE_DATE_ASC",
                "receipt_account": "一帆光大",
                "note": "客户统一付款，按默认优先分摊",
                "allocations": [
                    {"billing_record_id": record_a_id, "allocated_amount": 1800},
                    {"billing_record_id": record_b_id, "allocated_amount": 700},
                ],
            },
        )
        assert apply_resp.status_code == 201
        assert len(apply_resp.json()["allocations"]) == 2

        records_resp = client.get("/api/v1/billing-records", headers=headers)
        assert records_resp.status_code == 200
        records_data = records_resp.json()
        target_a = next(item for item in records_data if item["id"] == record_a_id)
        target_b = next(item for item in records_data if item["id"] == record_b_id)
        assert target_a["received_amount"] == 1800
        assert target_a["outstanding_amount"] == 200
        assert target_a["status"] == "PARTIAL"
        assert target_b["received_amount"] == 700
        assert target_b["outstanding_amount"] == 1100
        assert target_b["status"] == "PARTIAL"

        activities_a = client.get(f"/api/v1/billing-records/{record_a_id}/activities", headers=headers)
        assert activities_a.status_code == 200
        assert any("统一收款分摊" in (item["content"] or "") for item in activities_a.json())
        assert any(item["receipt_account"] == "一帆光大" for item in activities_a.json())


def test_renew_and_terminate_billing_record_lifecycle():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "续费终止测试客户",
                "contact_name": "生命周期",
                "phone": "13800138062",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "serial_no": 1601,
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "period_start_month": "2026-01",
                "period_end_month": "2026-12",
                "collection_start_date": "2026-01-01",
                "due_month": "2026-12-31",
                "total_fee": 2400,
                "monthly_fee": 200,
                "received_amount": 600,
                "payment_method": "后收",
                "summary": "原合同",
            },
        )
        assert record_resp.status_code == 201
        source_record_id = record_resp.json()["id"]

        renew_resp = client.post(
            f"/api/v1/billing-records/{source_record_id}/renew",
            headers=headers,
            json={"note": "自动续费生成"},
        )
        assert renew_resp.status_code == 201
        renewed = renew_resp.json()
        assert renewed["serial_no"] != 1601
        assert renewed["period_start_month"] == "2027-01"
        assert renewed["period_end_month"] == "2027-12"
        assert renewed["due_month"] == "2027-12-31"
        assert renewed["received_amount"] == 0

        terminate_resp = client.post(
            f"/api/v1/billing-records/{source_record_id}/terminate",
            headers=headers,
            json={
                "terminated_at": "2026-06-30",
                "reduced_fee": 800,
                "reason": "合同提前终止",
            },
        )
        assert terminate_resp.status_code == 200
        terminated = terminate_resp.json()
        assert terminated["total_fee"] == 1600
        assert terminated["due_month"] == "2026-06-30"
        assert terminated["period_end_month"] == "2026-06"
        assert "合同提前终止" in (terminated["note"] or "")


def test_renew_billing_record_supports_form_overrides():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "续费覆盖测试客户",
                "contact_name": "续费表单",
                "phone": "13800138102",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "charge_category": "代账",
                "amount_basis": "MONTHLY",
                "period_start_month": "2026-02",
                "period_end_month": "2027-01",
                "collection_start_date": "2026-02-01",
                "due_month": "2027-01-31",
                "total_fee": 3600,
                "monthly_fee": 300,
                "payment_method": "后收",
                "summary": "旧续费合同",
                "note": "原备注",
                "extra_note": "原扩展",
            },
        )
        assert record_resp.status_code == 201
        source_record_id = record_resp.json()["id"]

        renew_resp = client.post(
            f"/api/v1/billing-records/{source_record_id}/renew",
            headers=headers,
            json={
                "charge_category": "代账并退税",
                "summary": "续费后新合同",
                "total_fee": 4200,
                "monthly_fee": 350,
                "period_start_month": "2027-02",
                "period_end_month": "2028-01",
                "collection_start_date": "2027-02-01",
                "due_month": "2028-01-28",
                "payment_method": "预收",
                "status": "PARTIAL",
                "received_amount": 1200,
                "note": "续费后备注",
                "extra_note": "续费后扩展",
            },
        )
        assert renew_resp.status_code == 201
        renewed = renew_resp.json()
        assert renewed["charge_category"] == "代账并退税"
        assert renewed["summary"] == "续费后新合同"
        assert renewed["total_fee"] == 4200
        assert renewed["monthly_fee"] == 350
        assert renewed["period_start_month"] == "2027-02"
        assert renewed["period_end_month"] == "2028-01"
        assert renewed["collection_start_date"] == "2027-02-01"
        assert renewed["due_month"] == "2028-01-28"
        assert renewed["payment_method"] == "预收"
        assert renewed["received_amount"] == 1200
        assert renewed["outstanding_amount"] == 3000
        assert renewed["status"] == "PARTIAL"
        assert renewed["note"] == "续费后备注"
        assert renewed["extra_note"] == "续费后扩展"


def test_terminate_billing_record_rejects_date_outside_service_window():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "提前终止校验客户",
                "contact_name": "校验人",
                "phone": "13800138092",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "period_start_month": "2026-05",
                "period_end_month": "2026-12",
                "collection_start_date": "2026-05-01",
                "due_month": "2026-12-31",
                "total_fee": 2400,
                "payment_method": "后收",
                "summary": "终止日期校验合同",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        too_early_resp = client.post(
            f"/api/v1/billing-records/{record_id}/terminate",
            headers=headers,
            json={
                "terminated_at": "2026-04-30",
                "reduced_fee": 100,
                "reason": "错误日期",
            },
        )
        assert too_early_resp.status_code == 400
        assert too_early_resp.json()["detail"] == "终止日期不能早于服务开始日期"

        too_late_resp = client.post(
            f"/api/v1/billing-records/{record_id}/terminate",
            headers=headers,
            json={
                "terminated_at": "2027-01-01",
                "reduced_fee": 100,
                "reason": "错误日期",
            },
        )
        assert too_late_resp.status_code == 400
        assert too_late_resp.json()["detail"] == "终止日期不能晚于当前到期日"


def test_customer_billing_ledger_entries_and_balance():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "明细账测试客户",
                "contact_name": "台账负责人",
                "phone": "13800138063",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "serial_no": 1701,
                "customer_id": customer_id,
                "total_fee": 3000,
                "monthly_fee": 250,
                "due_month": "2026-03-31",
                "payment_method": "后收",
                "summary": "2026代账服务费",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        payment_resp = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": "2026-03-15",
                "amount": 1200,
                "payment_nature": "ONE_OFF",
                "receipt_account": "微信",
                "content": "客户首笔回款",
            },
        )
        assert payment_resp.status_code == 201
        payment_resp_2 = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": "2026-04-10",
                "amount": 600,
                "payment_nature": "ONE_OFF",
                "receipt_account": "支付宝",
                "content": "客户第二笔回款",
            },
        )
        assert payment_resp_2.status_code == 201

        ledger_resp = client.get(
            "/api/v1/billing-records/ledger",
            headers=headers,
            params={
                "customer_id": customer_id,
                "date_from": "2026-03-01",
                "date_to": "2026-04-30",
            },
        )
        assert ledger_resp.status_code == 200
        data = ledger_resp.json()
        assert data["customer_id"] == customer_id
        assert data["receivable_total"] == 3000
        assert data["received_total"] == 1800
        assert data["balance"] == 1200
        assert len(data["entries"]) >= 3
        assert any(item["source_type"] == "RECEIVABLE" for item in data["entries"])
        assert any(item["source_type"] == "PAYMENT" for item in data["entries"])
        assert any(item["receipt_account"] == "微信" for item in data["entries"] if item["source_type"] == "PAYMENT")
        assert any(item["receipt_account"] == "支付宝" for item in data["entries"] if item["source_type"] == "PAYMENT")
        assert data["entries"][-1]["balance"] == 1200
        assert len(data["monthly_summaries"]) >= 2
        march_summary = next(item for item in data["monthly_summaries"] if item["month"] == "2026-03")
        april_summary = next(item for item in data["monthly_summaries"] if item["month"] == "2026-04")
        assert march_summary["receivable_total"] == 3000
        assert march_summary["received_total"] == 1200
        assert march_summary["net_change"] == 1800
        assert march_summary["ending_balance"] == 1800
        assert april_summary["receivable_total"] == 0
        assert april_summary["received_total"] == 600
        assert april_summary["net_change"] == -600
        assert april_summary["ending_balance"] == 1200


def test_customer_update():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        customers = client.get("/api/v1/customers", headers=headers).json()
        customer_id = customers[0]["id"]
        update_resp = client.patch(
            f"/api/v1/customers/{customer_id}",
            headers=headers,
            json={
                "name": "可编辑客户名称",
                "contact_name": "新联系人",
                "phone": "13900000000",
                "lead_notes": "更新后的备注",
            },
        )
        assert update_resp.status_code == 200
        data = update_resp.json()
        assert data["name"] == "可编辑客户名称"
        assert data["contact_name"] == "新联系人"
        assert data["phone"] == "13900000000"
        assert data["lead"]["notes"] == "更新后的备注"


def test_receipt_account_ledger_for_admin():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={"name": "账户流水客户", "contact_name": "账务负责人", "phone": "13800138158"},
        )
        assert lead_resp.status_code == 201
        customer_id = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        ).json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "PERIODIC",
                "collection_start_date": "2026-03-01",
                "due_month": "2026-03-31",
                "total_fee": 3000,
                "payment_method": "后收",
                "summary": "2026年3月代账费",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        direct_payment = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": "2026-03-08",
                "amount": 1200,
                "payment_nature": "ONE_OFF",
                "receipt_account": "微信",
                "content": "第一笔到账",
            },
        )
        assert direct_payment.status_code == 201

        split_payment = client.post(
            "/api/v1/billing-records/payments",
            headers=headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-03-10",
                "amount": 1000,
                "strategy": "SERIAL_ASC",
                "receipt_account": "一帆光大",
                "note": "第二笔到账",
                "allocations": [
                    {"billing_record_id": record_id, "allocated_amount": 1000},
                ],
            },
        )
        assert split_payment.status_code == 201

        ledger_resp = client.get(
            "/api/v1/billing-records/receipt-account-ledger",
            headers=headers,
            params={"date_from": "2026-03-01", "date_to": "2026-03-31"},
        )
        assert ledger_resp.status_code == 200
        ledger_data = ledger_resp.json()
        assert ledger_data["payment_count"] >= 2
        assert ledger_data["total_received"] >= 2200
        assert any(item["receipt_account"] == "微信" for item in ledger_data["account_summaries"])
        assert any(item["receipt_account"] == "一帆光大" for item in ledger_data["account_summaries"])
        assert any(item["summary"] == "第二笔到账" for item in ledger_data["entries"])

        records_resp = client.get("/api/v1/billing-records", headers=headers)
        assert records_resp.status_code == 200
        record_row = next(item for item in records_resp.json() if item["id"] == record_id)
        assert record_row["receivable_period_text"] == "26.3-26.3"
        assert record_row["latest_receipt_account"] == "一帆光大"

        account_filtered_resp = client.get(
            "/api/v1/billing-records",
            headers=headers,
            params={"receipt_account": "一帆光大"},
        )
        assert account_filtered_resp.status_code == 200
        assert any(item["id"] == record_id for item in account_filtered_resp.json())

        customer_filtered_resp = client.get(
            "/api/v1/billing-records",
            headers=headers,
            params={"customer_id": customer_id},
        )
        assert customer_filtered_resp.status_code == 200
        assert any(item["id"] == record_id for item in customer_filtered_resp.json())

        summary_resp = client.get(
            "/api/v1/billing-records/summary",
            headers=headers,
            params={
                "customer_id": customer_id,
                "receipt_account": "一帆光大",
                "billing_month": "2026-03",
            },
        )
        assert summary_resp.status_code == 200
        summary_data = summary_resp.json()
        assert summary_data["total_records"] == 1
        assert summary_data["total_fee"] == 3000
        assert any(item["receipt_account"] == "一帆光大" for item in summary_data["receipt_account_distribution"])


def test_convert_reject_non_accountant_assignment():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        create_lead = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "FOLLOWUP",
                "name": "非会计分配测试",
                "contact_name": "王测试",
                "phone": "13800990099",
            },
        )
        assert create_lead.status_code == 201
        lead_id = create_lead.json()["id"]

        # admin 不是会计角色，应该被拒绝
        me_admin = client.post("/api/v1/auth/login", json={"username": "admin", "password": DEMO_PASSWORD})
        assert me_admin.status_code == 200
        admin_me = client.get(
            "/api/v1/auth/me",
            headers={"Authorization": f"Bearer {me_admin.json()['access_token']}"},
        )
        admin_id = admin_me.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": admin_id},
        )
        assert convert_resp.status_code == 400


def test_convert_requires_accountant_when_owner_not_accountant():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        create_lead = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "FOLLOWUP",
                "name": "必须选择会计测试",
                "contact_name": "王测试",
                "phone": "13800990100",
            },
        )
        assert create_lead.status_code == 201
        lead_id = create_lead.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={},
        )
        assert convert_resp.status_code == 400


def test_user_admin_scope_owner_vs_admin():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        owner_list = client.get("/api/v1/users", headers=owner_headers, params={"include_inactive": True})
        assert owner_list.status_code == 200
        assert all(item["role"] != "ADMIN" for item in owner_list.json())

        owner_create_accountant = client.post(
            "/api/v1/users",
            headers=owner_headers,
            json={
                "username": "owner_created_acc",
                "password": DEMO_PASSWORD,
                "role": "ACCOUNTANT",
                "is_active": True,
            },
        )
        assert owner_create_accountant.status_code == 201

        owner_create_admin = client.post(
            "/api/v1/users",
            headers=owner_headers,
            json={
                "username": "owner_created_admin",
                "password": DEMO_PASSWORD,
                "role": "ADMIN",
            },
        )
        assert owner_create_admin.status_code == 403

        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        admin_headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}
        admin_me = client.get("/api/v1/auth/me", headers=admin_headers)
        admin_id = admin_me.json()["id"]

        owner_patch_admin = client.patch(
            f"/api/v1/users/{admin_id}",
            headers=owner_headers,
            json={"is_active": False},
        )
        assert owner_patch_admin.status_code == 403

        admin_create_admin = client.post(
            "/api/v1/users",
            headers=admin_headers,
            json={
                "username": "admin_created_admin",
                "password": DEMO_PASSWORD,
                "role": "ADMIN",
            },
        )
        assert admin_create_admin.status_code == 201


def test_user_update_self_protection():
    with TestClient(app) as client:
        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}
        me_resp = client.get("/api/v1/auth/me", headers=headers)
        me_id = me_resp.json()["id"]

        deactivate_self = client.patch(
            f"/api/v1/users/{me_id}",
            headers=headers,
            json={"is_active": False},
        )
        assert deactivate_self.status_code == 400

        change_self_role = client.patch(
            f"/api/v1/users/{me_id}",
            headers=headers,
            json={"role": "ACCOUNTANT"},
        )
        assert change_self_role.status_code == 400


def test_user_delete_flow_and_dependency_guard():
    with TestClient(app) as client:
        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}

        create_response = client.post(
            "/api/v1/users",
            headers=headers,
            json={
                "username": "delete_me",
                "password": DEMO_PASSWORD,
                "role": "ACCOUNTANT",
                "is_active": True,
            },
        )
        assert create_response.status_code == 201
        delete_id = create_response.json()["id"]

        delete_response = client.delete(
            f"/api/v1/users/{delete_id}",
            headers=headers,
            params={"confirm_name": "delete_me"},
        )
        assert delete_response.status_code == 204

        list_response = client.get("/api/v1/users", headers=headers, params={"include_inactive": True})
        assert list_response.status_code == 200
        assert all(item["id"] != delete_id for item in list_response.json())

        login_deleted = client.post(
            "/api/v1/auth/login",
            json={"username": "delete_me", "password": DEMO_PASSWORD},
        )
        assert login_deleted.status_code == 401

        create_bound_user = client.post(
            "/api/v1/users",
            headers=headers,
            json={
                "username": "bound_acc",
                "password": DEMO_PASSWORD,
                "role": "ACCOUNTANT",
                "is_active": True,
            },
        )
        assert create_bound_user.status_code == 201
        bound_user_id = create_bound_user.json()["id"]

        bound_login = client.post(
            "/api/v1/auth/login",
            json={"username": "bound_acc", "password": DEMO_PASSWORD},
        )
        assert bound_login.status_code == 200
        bound_headers = {"Authorization": f"Bearer {bound_login.json()['access_token']}"}
        bound_lead = client.post(
            "/api/v1/leads",
            headers=bound_headers,
            json={
                "template_type": "FOLLOWUP",
                "name": "删除阻断测试线索",
                "contact_name": "联系人",
                "phone": "13800112233",
            },
        )
        assert bound_lead.status_code == 201

        blocked_delete = client.delete(f"/api/v1/users/{bound_user_id}", headers=headers)
        assert blocked_delete.status_code == 400
        assert "关联数据" in blocked_delete.json()["detail"]


def test_login_updates_last_login_and_operation_log():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        token = login_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        me_response = client.get("/api/v1/auth/me", headers=headers)
        assert me_response.status_code == 200
        assert me_response.json()["last_login_at"] is not None

        logs_response = client.get(
            "/api/v1/admin/operation-logs",
            headers=headers,
            params={"action": "LOGIN", "limit": 50},
        )
        assert logs_response.status_code == 200
        assert any(
            item["action"] == "LOGIN" and item["actor_username"] == "boss"
            for item in logs_response.json()
        )


def test_ldap_settings_permission_and_sync_guard():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        admin_headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}

        owner_set_admin_default = client.put(
            "/api/v1/admin/ldap/settings",
            headers=owner_headers,
            json={"default_role": "ADMIN"},
        )
        assert owner_set_admin_default.status_code == 403

        admin_update = client.put(
            "/api/v1/admin/ldap/settings",
            headers=admin_headers,
            json={
                "enabled": False,
                "server_url": "ldap://127.0.0.1:389",
                "bind_dn": "uid=admin,cn=users,dc=example,dc=com",
                "base_dn": "dc=example,dc=com",
                "user_base_dn": "cn=users,dc=example,dc=com",
                "user_filter": "(uid=*)",
                "username_attr": "uid",
                "display_name_attr": "cn",
                "default_role": "ACCOUNTANT",
            },
        )
        assert admin_update.status_code == 200
        assert admin_update.json()["default_role"] == "ACCOUNTANT"

        admin_update_manager = client.put(
            "/api/v1/admin/ldap/settings",
            headers=admin_headers,
            json={"default_role": "MANAGER"},
        )
        assert admin_update_manager.status_code == 200
        assert admin_update_manager.json()["default_role"] == "MANAGER"

        sync_response = client.post("/api/v1/admin/ldap/sync", headers=admin_headers, json={})
        assert sync_response.status_code == 400


def test_dashboard_and_todo_flow():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        users = users_resp.json()
        accountant2 = next(item for item in users if item["username"] == "accountant2")

        create_for_other = client.post(
            "/api/v1/todos",
            headers=owner_headers,
            json={
                "title": "月底工资核算复核",
                "description": "仅查看授权数据，不可写入",
                "priority": "HIGH",
                "due_date": "2026-03-05",
                "assignee_user_id": accountant2["id"],
            },
        )
        assert create_for_other.status_code == 201
        todo_id = create_for_other.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant2", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        list_open_todos = client.get("/api/v1/todos", headers=accountant_headers)
        assert list_open_todos.status_code == 200
        assert any(item["id"] == todo_id for item in list_open_todos.json())

        mark_done = client.patch(
            f"/api/v1/todos/{todo_id}",
            headers=accountant_headers,
            json={"status": "DONE"},
        )
        assert mark_done.status_code == 200
        assert mark_done.json()["status"] == "DONE"

        dashboard_summary = client.get("/api/v1/dashboard/summary", headers=accountant_headers)
        assert dashboard_summary.status_code == 200
        summary_data = dashboard_summary.json()
        assert summary_data["month"]
        assert "system_todo_count" in summary_data
        assert "manual_open_todo_count" in summary_data

        system_todos = client.get("/api/v1/dashboard/system-todos", headers=accountant_headers, params={"limit": 20})
        assert system_todos.status_code == 200
        assert isinstance(system_todos.json(), list)


def test_todo_delete_requires_admin_and_confirm_name():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        create_resp = client.post(
            "/api/v1/todos",
            headers=owner_headers,
            json={
                "title": "待办删除测试-A",
                "priority": "HIGH",
                "assignee_user_id": accountant["id"],
            },
        )
        assert create_resp.status_code == 201
        todo_id = create_resp.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        forbidden_resp = client.delete(
            f"/api/v1/todos/{todo_id}",
            headers=accountant_headers,
            params={"confirm_name": "待办删除测试-A"},
        )
        assert forbidden_resp.status_code == 403

        mismatch_resp = client.delete(
            f"/api/v1/todos/{todo_id}",
            headers=owner_headers,
            params={"confirm_name": "错误名称"},
        )
        assert mismatch_resp.status_code == 400

        delete_resp = client.delete(
            f"/api/v1/todos/{todo_id}",
            headers=owner_headers,
            params={"confirm_name": "待办删除测试-A"},
        )
        assert delete_resp.status_code == 204

        list_resp = client.get(
            "/api/v1/todos",
            headers=owner_headers,
            params={"view": "ALL", "include_done": True, "assignee_user_id": accountant["id"]},
        )
        assert list_resp.status_code == 200
        assert all(item["id"] != todo_id for item in list_resp.json())

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "TODO"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == todo_id for item in deleted_rows.json())

        restore_resp = client.post(
            f"/api/v1/admin/deleted-records/TODO/{todo_id}/restore",
            headers=owner_headers,
        )
        assert restore_resp.status_code == 200
        assert restore_resp.json()["entity_type"] == "TODO"

        restored_list_resp = client.get(
            "/api/v1/todos",
            headers=owner_headers,
            params={"view": "ALL", "include_done": True, "assignee_user_id": accountant["id"]},
        )
        assert restored_list_resp.status_code == 200
        assert any(item["id"] == todo_id for item in restored_list_resp.json())

        restore_logs = client.get(
            "/api/v1/admin/operation-logs",
            headers=owner_headers,
            params={"audit_scope": "RESTORE", "entity_type": "TODO"},
        )
        assert restore_logs.status_code == 200
        assert any(item["action"] == "TODO_RESTORED" for item in restore_logs.json())


def test_lead_delete_uses_simple_confirm_and_blocks_converted():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "FOLLOWUP",
                "name": "线索删除测试-A",
                "contact_name": "李女士",
                "phone": "13950001001",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        forbidden_resp = client.delete(
            f"/api/v1/leads/{lead_id}",
            headers=accountant_headers,
        )
        assert forbidden_resp.status_code == 403

        delete_resp = client.delete(
            f"/api/v1/leads/{lead_id}",
            headers=owner_headers,
        )
        assert delete_resp.status_code == 204

        missing_resp = client.get(f"/api/v1/leads/{lead_id}", headers=owner_headers)
        assert missing_resp.status_code == 404

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "LEAD"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == lead_id for item in deleted_rows.json())

        restore_resp = client.post(
            f"/api/v1/admin/deleted-records/LEAD/{lead_id}/restore",
            headers=owner_headers,
        )
        assert restore_resp.status_code == 200

        restored_detail = client.get(f"/api/v1/leads/{lead_id}", headers=owner_headers)
        assert restored_detail.status_code == 200
        assert restored_detail.json()["name"] == "线索删除测试-A"

        converted_lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "线索删除测试-B",
                "contact_name": "王总",
                "phone": "13950001002",
                "source": "Sally直播",
            },
        )
        assert converted_lead_resp.status_code == 201
        converted_lead_id = converted_lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{converted_lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200

        converted_delete_resp = client.delete(
            f"/api/v1/leads/{converted_lead_id}",
            headers=owner_headers,
        )
        assert converted_delete_resp.status_code == 400
        assert converted_delete_resp.json()["detail"] == "该线索已转化，请先撤销转化再删除"


def test_customer_delete_requires_admin_confirm_and_supports_restore():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "客户删除测试-A",
                "contact_name": "赵小姐",
                "phone": "13950002001",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        redevelop_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "REDEVELOP",
                "name": "客户删除测试-关联线索",
                "contact_name": "赵小姐",
                "phone": "13950002001",
                "source": "老客户二次开发",
                "related_customer_id": customer_id,
            },
        )
        assert redevelop_resp.status_code == 201
        redevelop_lead_id = redevelop_resp.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        forbidden_resp = client.delete(
            f"/api/v1/customers/{customer_id}",
            headers=accountant_headers,
            params={"confirm_name": "客户删除测试-A"},
        )
        assert forbidden_resp.status_code == 403

        mismatch_resp = client.delete(
            f"/api/v1/customers/{customer_id}",
            headers=owner_headers,
            params={"confirm_name": "错误名称"},
        )
        assert mismatch_resp.status_code == 400

        delete_resp = client.delete(
            f"/api/v1/customers/{customer_id}",
            headers=owner_headers,
            params={"confirm_name": "客户删除测试-A"},
        )
        assert delete_resp.status_code == 204

        customer_detail = client.get(f"/api/v1/customers/{customer_id}", headers=owner_headers)
        assert customer_detail.status_code == 404

        lead_detail = client.get(f"/api/v1/leads/{lead_id}", headers=owner_headers)
        assert lead_detail.status_code == 200
        assert lead_detail.json()["status"] == "CONVERTED"

        redevelop_detail = client.get(f"/api/v1/leads/{redevelop_lead_id}", headers=owner_headers)
        assert redevelop_detail.status_code == 200
        assert redevelop_detail.json()["related_customer_id"] == customer_id

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "CUSTOMER"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == customer_id for item in deleted_rows.json())

        restore_resp = client.post(
            f"/api/v1/admin/deleted-records/CUSTOMER/{customer_id}/restore",
            headers=owner_headers,
        )
        assert restore_resp.status_code == 200

        restored_customer_detail = client.get(f"/api/v1/customers/{customer_id}", headers=owner_headers)
        assert restored_customer_detail.status_code == 200
        assert restored_customer_detail.json()["name"] == "客户删除测试-A"


def test_unconvert_then_reconvert_reuses_soft_deleted_customer():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "反复成交测试-A",
                "contact_name": "林老板",
                "phone": "13950002011",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        first_convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert first_convert_resp.status_code == 200
        customer_id = first_convert_resp.json()["customer"]["id"]

        unconvert_resp = client.post(f"/api/v1/leads/{lead_id}/unconvert", headers=owner_headers)
        assert unconvert_resp.status_code == 200
        assert unconvert_resp.json()["status"] == "NEW"

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "CUSTOMER"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == customer_id for item in deleted_rows.json())

        second_convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={
                "accountant_id": accountant["id"],
                "customer_name": "反复成交测试-A（再次成交）",
            },
        )
        assert second_convert_resp.status_code == 200
        assert second_convert_resp.json()["customer"]["id"] == customer_id
        assert second_convert_resp.json()["customer"]["name"] == "反复成交测试-A（再次成交）"


def test_customer_delete_returns_structured_blockers():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "删除阻塞客户",
                "contact_name": "阻塞联系人",
                "phone": "13950002101",
                "source": "Sally直播",
                "main_business": "代账服务",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "删除阻塞收费项目",
                "total_fee": 600,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-07",
                "due_month": "2026-04-07",
                "payment_method": "后收",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        payment_resp = client.post(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-04-07",
                "amount": 600,
                "strategy": "DUE_DATE_ASC",
                "receipt_account": "一帆光大",
                "note": "删除阻塞收款单",
                "allocations": [{"billing_record_id": record_id, "allocated_amount": 600}],
            },
        )
        assert payment_resp.status_code == 201

        matter_resp = client.post(
            f"/api/v1/customers/{customer_id}/timeline-events",
            headers=owner_headers,
            json={
                "occurred_at": "2026-04-07",
                "event_type": "DELIVERY",
                "status": "OPEN",
                "reminder_at": "2026-04-10",
                "content": "客户需要补充资料",
                "note": "删除前不应允许丢失事项",
                "result": "",
                "amount": None,
            },
        )
        assert matter_resp.status_code == 201

        delete_resp = client.delete(
            f"/api/v1/customers/{customer_id}",
            headers=owner_headers,
            params={"confirm_name": "删除阻塞客户"},
        )
        assert delete_resp.status_code == 400
        detail = delete_resp.json()["detail"]
        assert detail["reason"] == "DEPENDENCY_BLOCKED"
        blocker_types = {item["type"] for item in detail["blockers"]}
        assert "BILLING_RECORD" in blocker_types
        assert "BILLING_PAYMENT" in blocker_types
        assert "CUSTOMER_MATTER" in blocker_types


def test_lead_convert_generates_customer_code_and_suffix():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "编号客户-S",
                "contact_name": "编号联系人",
                "phone": "13950002102",
                "source": "Sally直播",
                "main_business": "咨询",
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer = convert_resp.json()["customer"]
        assert customer["customer_code_seq"] is not None
        assert customer["customer_code_suffix"] == "S"
        assert customer["customer_code"].endswith("S")

        list_resp = client.get("/api/v1/customers", headers=owner_headers, params={"keyword": customer["customer_code"]})
        assert list_resp.status_code == 200
        assert any(item["id"] == customer["id"] for item in list_resp.json())


def test_billing_record_delete_uses_simple_confirm_and_blocks_paid_record():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "收费删除测试-A",
                "contact_name": "钱先生",
                "phone": "13950003001",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "咨询",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "收费删除测试-未收",
                "total_fee": 800,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-06",
                "due_month": "2026-04-06",
                "payment_method": "后收",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        forbidden_resp = client.delete(
            f"/api/v1/billing-records/{record_id}",
            headers=accountant_headers,
        )
        assert forbidden_resp.status_code == 403

        delete_resp = client.delete(
            f"/api/v1/billing-records/{record_id}",
            headers=owner_headers,
        )
        assert delete_resp.status_code == 204

        list_resp = client.get("/api/v1/billing-records", headers=owner_headers)
        assert list_resp.status_code == 200
        assert all(item["id"] != record_id for item in list_resp.json())

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "BILLING"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == record_id for item in deleted_rows.json())

        restore_resp = client.post(
            f"/api/v1/admin/deleted-records/BILLING/{record_id}/restore",
            headers=owner_headers,
        )
        assert restore_resp.status_code == 200

        restored_list_resp = client.get("/api/v1/billing-records", headers=owner_headers)
        assert restored_list_resp.status_code == 200
        assert any(item["id"] == record_id for item in restored_list_resp.json())

        paid_record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "注册",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "收费删除测试-已收",
                "total_fee": 1200,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-07",
                "due_month": "2026-04-07",
                "payment_method": "后收",
            },
        )
        assert paid_record_resp.status_code == 201
        paid_record_id = paid_record_resp.json()["id"]

        payment_resp = client.post(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-04-07",
                "amount": 1200,
                "strategy": "DUE_DATE_ASC",
                "receipt_account": "一帆光大",
                "note": "删除校验测试",
                "allocations": [
                    {"billing_record_id": paid_record_id, "allocated_amount": 1200},
                ],
            },
        )
        assert payment_resp.status_code == 201

        paid_delete_resp = client.delete(
            f"/api/v1/billing-records/{paid_record_id}",
            headers=owner_headers,
        )
        assert paid_delete_resp.status_code == 400
        detail = paid_delete_resp.json()["detail"]
        assert detail["reason"] == "DEPENDENCY_BLOCKED"
        assert detail["blockers"][0]["type"] == "BILLING_PAYMENT"
        assert detail["blockers"][0]["filters"]["recordId"] == paid_record_id
        assert detail["blockers"][0]["filters"]["focusDependency"] == 1
        assert f"recordId={paid_record_id}" in detail["blockers"][0]["href"]


def test_billing_payment_supports_prepay_list_and_manual_allocate():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "template_type": "CONVERSION",
                "name": "预收款测试客户",
                "contact_name": "预收联系人",
                "phone": "13950003111",
                "source": "Sally直播",
                "main_business": "代账",
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        plain_payment_resp = client.post(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-04-07",
                "amount": 500,
                "strategy": "DUE_DATE_ASC",
                "receipt_account": "一帆光大",
                "note": "无应收普通收款",
                "allocations": [],
            },
        )
        assert plain_payment_resp.status_code == 400
        assert plain_payment_resp.json()["detail"]["reason"] == "NO_ALLOCATIONS"

        prepay_resp = client.post(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-04-07",
                "amount": 500,
                "strategy": "DUE_DATE_ASC",
                "receipt_account": "一帆光大",
                "note": "客户先付预收款",
                "is_prepay": True,
                "allocations": [],
            },
        )
        assert prepay_resp.status_code == 201
        payment_id = prepay_resp.json()["id"]
        assert prepay_resp.json()["allocation_status"] == "UNALLOCATED"
        assert prepay_resp.json()["unallocated_amount"] == 500

        list_resp = client.get(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            params={"customer_id": customer_id, "unallocated_only": True},
        )
        assert list_resp.status_code == 200
        assert any(item["id"] == payment_id for item in list_resp.json())

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "后续形成应收",
                "total_fee": 500,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-08",
                "due_month": "2026-04-08",
                "payment_method": "后收",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        allocate_resp = client.post(
            f"/api/v1/billing-records/payments/{payment_id}/allocate",
            headers=owner_headers,
            json={"allocations": [{"billing_record_id": record_id, "allocated_amount": 500}]},
        )
        assert allocate_resp.status_code == 200
        assert allocate_resp.json()["allocation_status"] == "ALLOCATED"
        assert allocate_resp.json()["unallocated_amount"] == 0

        records_resp = client.get("/api/v1/billing-records", headers=owner_headers, params={"customer_id": customer_id})
        assert records_resp.status_code == 200
        target = next(item for item in records_resp.json() if item["id"] == record_id)
        assert target["received_amount"] == 500
        assert target["status"] == "CLEARED"


def test_billing_payment_list_supports_record_filter():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "收款列表按收费项目过滤客户",
                "contact_name": "过滤联系人",
                "phone": "13800131234",
                "main_business": "收费过滤测试",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_a_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "收费项目A",
                "total_fee": 600,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-08",
                "due_month": "2026-04-08",
                "payment_method": "后收",
            },
        )
        assert record_a_resp.status_code == 201
        record_a_id = record_a_resp.json()["id"]

        record_b_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "注册",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "收费项目B",
                "total_fee": 300,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-09",
                "due_month": "2026-04-09",
                "payment_method": "后收",
            },
        )
        assert record_b_resp.status_code == 201
        record_b_id = record_b_resp.json()["id"]

        payment_resp = client.post(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "occurred_at": "2026-04-09",
                "amount": 600,
                "strategy": "DUE_DATE_ASC",
                "receipt_account": "一帆光大",
                "note": "仅核销收费项目A",
                "allocations": [{"billing_record_id": record_a_id, "allocated_amount": 600}],
            },
        )
        assert payment_resp.status_code == 201
        payment_id = payment_resp.json()["id"]

        record_filtered_resp = client.get(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            params={"record_id": record_a_id},
        )
        assert record_filtered_resp.status_code == 200
        filtered_rows = record_filtered_resp.json()
        assert [item["id"] for item in filtered_rows] == [payment_id]

        empty_record_resp = client.get(
            "/api/v1/billing-records/payments",
            headers=owner_headers,
            params={"record_id": record_b_id},
        )
        assert empty_record_resp.status_code == 200
        assert empty_record_resp.json() == []


def test_billing_record_delete_blocks_direct_payment_activity_with_activity_target():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "旧版收费活动收款删除校验客户",
                "contact_name": "旧版活动联系人",
                "phone": "13800131235",
                "main_business": "删除校验",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "旧版活动收款收费项目",
                "total_fee": 300,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-09",
                "due_month": "2026-04-09",
                "payment_method": "后收",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        activity_resp = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=owner_headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": "2026-04-09",
                "amount": 300,
                "payment_nature": "ONE_OFF",
                "receipt_account": "一帆光大",
                "content": "旧版活动收款",
                "note": "没有收款单 allocation",
            },
        )
        assert activity_resp.status_code == 201
        activity_id = activity_resp.json()["id"]

        blocked_delete_resp = client.delete(f"/api/v1/billing-records/{record_id}", headers=owner_headers)
        assert blocked_delete_resp.status_code == 400
        detail = blocked_delete_resp.json()["detail"]
        assert detail["reason"] == "DEPENDENCY_BLOCKED"
        assert detail["blockers"][0]["type"] == "BILLING_ACTIVITY_PAYMENT"
        assert detail["blockers"][0]["filters"]["record_id"] == record_id
        assert detail["blockers"][0]["filters"]["action"] == "activity"
        assert detail["blockers"][0]["filters"]["focusDependency"] == 1

        remove_activity_resp = client.delete(
            f"/api/v1/billing-records/{record_id}/activities/{activity_id}",
            headers=owner_headers,
        )
        assert remove_activity_resp.status_code == 204

        records_resp = client.get("/api/v1/billing-records", headers=owner_headers, params={"customer_id": customer_id})
        assert records_resp.status_code == 200
        target = next(item for item in records_resp.json() if item["id"] == record_id)
        assert target["received_amount"] == 0
        assert target["outstanding_amount"] == 300

        final_delete_resp = client.delete(f"/api/v1/billing-records/{record_id}", headers=owner_headers)
        assert final_delete_resp.status_code == 204


def test_billing_record_delete_recovers_from_stale_received_amount_without_real_dependencies():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "脏实收字段删除校验客户",
                "contact_name": "脏数据联系人",
                "phone": "13800131236",
                "main_business": "脏值校验",
                "source": "Sally直播",
            },
        )
        assert lead_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_resp.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "ONE_TIME",
                "amount_basis": "ONE_TIME",
                "summary": "脏实收字段收费项目",
                "total_fee": 260,
                "monthly_fee": 0,
                "collection_start_date": "2026-04-09",
                "due_month": "2026-04-09",
                "payment_method": "后收",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        with SessionLocal() as db:
            record = db.get(BillingRecord, record_id)
            assert record is not None
            record.received_amount = 260
            db.commit()

        delete_resp = client.delete(f"/api/v1/billing-records/{record_id}", headers=owner_headers)
        assert delete_resp.status_code == 204


def test_user_delete_requires_confirm_and_supports_restore():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        username = f"qa_user_{uuid4().hex[:8]}"
        create_resp = client.post(
            "/api/v1/users",
            headers=owner_headers,
            json={
                "username": username,
                "password": DEMO_PASSWORD,
                "role": "ACCOUNTANT",
                "is_active": True,
            },
        )
        assert create_resp.status_code == 201
        user_id = create_resp.json()["id"]

        mismatch_resp = client.delete(
            f"/api/v1/users/{user_id}",
            headers=owner_headers,
            params={"confirm_name": "错误账号"},
        )
        assert mismatch_resp.status_code == 400

        delete_resp = client.delete(
            f"/api/v1/users/{user_id}",
            headers=owner_headers,
            params={"confirm_name": username},
        )
        assert delete_resp.status_code == 204

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"include_inactive": True})
        assert users_resp.status_code == 200
        assert all(item["id"] != user_id for item in users_resp.json())

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "USER"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == user_id for item in deleted_rows.json())

        deleted_login = client.post(
            "/api/v1/auth/login",
            json={"username": username, "password": DEMO_PASSWORD},
        )
        assert deleted_login.status_code == 401

        restore_resp = client.post(
            f"/api/v1/admin/deleted-records/USER/{user_id}/restore",
            headers=owner_headers,
        )
        assert restore_resp.status_code == 200
        assert restore_resp.json()["entity_type"] == "USER"

        restored_users = client.get("/api/v1/users", headers=owner_headers, params={"include_inactive": True})
        assert restored_users.status_code == 200
        assert any(item["id"] == user_id for item in restored_users.json())

        restored_login = client.post(
            "/api/v1/auth/login",
            json={"username": username, "password": DEMO_PASSWORD},
        )
        assert restored_login.status_code == 200

        restore_logs = client.get(
            "/api/v1/admin/operation-logs",
            headers=owner_headers,
            params={"audit_scope": "RESTORE", "entity_type": "USER"},
        )
        assert restore_logs.status_code == 200
        assert any(item["action"] == "USER_RESTORED" for item in restore_logs.json())


def test_data_access_grant_delete_requires_confirm_and_supports_restore():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        username = f"qa_grant_{uuid4().hex[:8]}"
        create_user_resp = client.post(
            "/api/v1/users",
            headers=owner_headers,
            json={
                "username": username,
                "password": DEMO_PASSWORD,
                "role": "ACCOUNTANT",
                "is_active": True,
            },
        )
        assert create_user_resp.status_code == 201
        user_id = create_user_resp.json()["id"]

        create_grant_resp = client.post(
            "/api/v1/admin/data-access-grants",
            headers=owner_headers,
            json={
                "grantee_user_id": user_id,
                "module": "CUSTOMER",
                "reason": "删除恢复测试",
                "is_active": True,
            },
        )
        assert create_grant_resp.status_code == 201
        grant_id = create_grant_resp.json()["id"]

        mismatch_resp = client.delete(
            f"/api/v1/admin/data-access-grants/{grant_id}",
            headers=owner_headers,
            params={"confirm_name": "错误名称"},
        )
        assert mismatch_resp.status_code == 400

        delete_resp = client.delete(
            f"/api/v1/admin/data-access-grants/{grant_id}",
            headers=owner_headers,
            params={"confirm_name": f"{username} · 客户列表"},
        )
        assert delete_resp.status_code == 204

        active_grants_resp = client.get("/api/v1/admin/data-access-grants", headers=owner_headers)
        assert active_grants_resp.status_code == 200
        assert all(item["id"] != grant_id for item in active_grants_resp.json())

        deleted_rows = client.get(
            "/api/v1/admin/deleted-records",
            headers=owner_headers,
            params={"entity_type": "DATA_ACCESS_GRANT"},
        )
        assert deleted_rows.status_code == 200
        assert any(item["entity_id"] == grant_id for item in deleted_rows.json())

        restore_resp = client.post(
            f"/api/v1/admin/deleted-records/DATA_ACCESS_GRANT/{grant_id}/restore",
            headers=owner_headers,
        )
        assert restore_resp.status_code == 200
        assert restore_resp.json()["entity_type"] == "DATA_ACCESS_GRANT"

        restored_grants_resp = client.get("/api/v1/admin/data-access-grants", headers=owner_headers)
        assert restored_grants_resp.status_code == 200
        assert any(item["id"] == grant_id for item in restored_grants_resp.json())

        restore_logs = client.get(
            "/api/v1/admin/operation-logs",
            headers=owner_headers,
            params={"audit_scope": "RESTORE", "entity_type": "DATA_ACCESS_GRANT"},
        )
        assert restore_logs.status_code == 200
        assert any(item["action"] == "DATA_ACCESS_GRANT_RESTORED" for item in restore_logs.json())


def test_billing_due_system_todo_lifecycle():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "到期待办闭环客户",
                "contact_name": "财务负责人",
                "phone": "13800138999",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        due_date = date.today() + timedelta(days=2)
        create_record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1901,
                "customer_id": customer_id,
                "total_fee": 3000,
                "monthly_fee": 300,
                "billing_cycle_text": "闭环测试账期",
                "due_month": due_date.isoformat(),
                "payment_method": "后收",
            },
        )
        assert create_record_resp.status_code == 201
        record_id = create_record_resp.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        system_todos_before = client.get("/api/v1/dashboard/system-todos", headers=accountant_headers, params={"limit": 200})
        assert system_todos_before.status_code == 200
        todo_ids_before = {item["id"] for item in system_todos_before.json()}
        assert f"billing:{record_id}" in todo_ids_before

        reminder_resp = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=accountant_headers,
            json={
                "activity_type": "REMINDER",
                "occurred_at": date.today().isoformat(),
                "amount": 0,
                "content": "电话催收，客户确认本周付款",
            },
        )
        assert reminder_resp.status_code == 201

        payment_resp = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=accountant_headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": date.today().isoformat(),
                "amount": 3000,
                "payment_nature": "YEARLY",
                "receipt_account": "一帆青岛",
                "is_settlement": True,
                "content": "款项到账，结清",
            },
        )
        assert payment_resp.status_code == 201

        records_resp = client.get("/api/v1/billing-records", headers=accountant_headers)
        assert records_resp.status_code == 200
        target_record = next(item for item in records_resp.json() if item["id"] == record_id)
        assert target_record["status"] == "CLEARED"
        assert target_record["outstanding_amount"] == 0

        system_todos_after = client.get("/api/v1/dashboard/system-todos", headers=accountant_headers, params={"limit": 200})
        assert system_todos_after.status_code == 200
        todo_ids_after = {item["id"] for item in system_todos_after.json()}
        assert f"billing:{record_id}" not in todo_ids_after


def test_billing_renew_system_todo_contains_action_path():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = next(item for item in users_resp.json() if item["username"] == "accountant")

        lead_resp = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={
                "name": "续费待办测试客户",
                "contact_name": "续费联系人",
                "phone": "13800138998",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        due_date = date.today() + timedelta(days=10)
        due_month_text = due_date.isoformat()[:7]
        create_record_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "serial_no": 1902,
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "period_start_month": "2026-01",
                "period_end_month": due_month_text,
                "due_month": due_date.isoformat(),
                "total_fee": 4800,
                "monthly_fee": 400,
                "payment_method": "预收",
            },
        )
        assert create_record_resp.status_code == 201
        record_id = create_record_resp.json()["id"]

        accountant_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert accountant_login.status_code == 200
        accountant_headers = {"Authorization": f"Bearer {accountant_login.json()['access_token']}"}

        system_todos = client.get("/api/v1/dashboard/system-todos", headers=accountant_headers, params={"limit": 200})
        assert system_todos.status_code == 200
        renew_todo = next(item for item in system_todos.json() if item["id"] == f"renew:{record_id}")
        assert renew_todo["action_label"] == "确认续费"
        assert renew_todo["action_path"] == f"/billing?action=renew&record_id={record_id}"


def test_system_billing_todo_only_for_assigned_accountant():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=owner_headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        users = users_resp.json()
        accountant = next(item for item in users if item["username"] == "accountant")
        accountant2 = next(item for item in users if item["username"] == "accountant2")

        lead_a = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={"name": "负责人A", "contact_name": "A", "phone": "13800138071"},
        )
        assert lead_a.status_code == 201
        customer_a = client.post(
            f"/api/v1/leads/{lead_a.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant["id"]},
        ).json()["customer"]["id"]

        lead_b = client.post(
            "/api/v1/leads",
            headers=owner_headers,
            json={"name": "负责人B", "contact_name": "B", "phone": "13800138072"},
        )
        assert lead_b.status_code == 201
        customer_b = client.post(
            f"/api/v1/leads/{lead_b.json()['id']}/convert",
            headers=owner_headers,
            json={"accountant_id": accountant2["id"]},
        ).json()["customer"]["id"]

        due_date = (date.today() + timedelta(days=1)).isoformat()
        record_a_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_a,
                "total_fee": 1200,
                "monthly_fee": 100,
                "billing_cycle_text": "负责人A账期",
                "due_month": due_date,
                "payment_method": "后收",
            },
        )
        assert record_a_resp.status_code == 201
        record_a_id = record_a_resp.json()["id"]

        record_b_resp = client.post(
            "/api/v1/billing-records",
            headers=owner_headers,
            json={
                "customer_id": customer_b,
                "total_fee": 1300,
                "monthly_fee": 100,
                "billing_cycle_text": "负责人B账期",
                "due_month": due_date,
                "payment_method": "后收",
            },
        )
        assert record_b_resp.status_code == 201
        record_b_id = record_b_resp.json()["id"]

        grant_resp = client.post(
            "/api/v1/admin/data-access-grants",
            headers=owner_headers,
            json={
                "grantee_user_id": accountant2["id"],
                "module": "BILLING",
                "reason": "临时查看全量收费",
            },
        )
        assert grant_resp.status_code == 201

        accountant2_login = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant2", "password": DEMO_PASSWORD},
        )
        assert accountant2_login.status_code == 200
        accountant2_headers = {"Authorization": f"Bearer {accountant2_login.json()['access_token']}"}

        accountant2_me = client.get("/api/v1/auth/me", headers=accountant2_headers)
        assert accountant2_me.status_code == 200
        assert "BILLING" in accountant2_me.json()["granted_read_modules"]

        billing_visible = client.get("/api/v1/billing-records", headers=accountant2_headers)
        assert billing_visible.status_code == 200
        visible_ids = {item["id"] for item in billing_visible.json()}
        assert record_a_id in visible_ids and record_b_id in visible_ids

        receipt_ledger_visible = client.get(
            "/api/v1/billing-records/receipt-account-ledger",
            headers=accountant2_headers,
            params={"date_from": "2026-03-01", "date_to": "2026-03-31"},
        )
        assert receipt_ledger_visible.status_code == 200

        todo_visible = client.get("/api/v1/dashboard/system-todos", headers=accountant2_headers, params={"limit": 200})
        assert todo_visible.status_code == 200
        todo_ids = {item["id"] for item in todo_visible.json()}
        assert f"billing:{record_b_id}" in todo_ids
        assert f"billing:{record_a_id}" not in todo_ids


def test_todo_my_day_toggle_and_carry_behavior():
    with TestClient(app) as client:
        login_response = client.post(
            "/api/v1/auth/login",
            json={"username": "accountant", "password": DEMO_PASSWORD},
        )
        assert login_response.status_code == 200
        headers = {"Authorization": f"Bearer {login_response.json()['access_token']}"}

        first_todo = client.post(
            "/api/v1/todos",
            headers=headers,
            json={
                "title": "今日机制测试A",
                "priority": "MEDIUM",
            },
        )
        assert first_todo.status_code == 201
        first_todo_id = first_todo.json()["id"]

        second_todo = client.post(
            "/api/v1/todos",
            headers=headers,
            json={
                "title": "今日机制测试B",
                "priority": "LOW",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
            },
        )
        assert second_todo.status_code == 201
        second_todo_id = second_todo.json()["id"]

        before_today = client.get("/api/v1/todos", headers=headers, params={"view": "TODAY"})
        assert before_today.status_code == 200
        before_ids = {item["id"] for item in before_today.json()}
        assert first_todo_id not in before_ids
        assert second_todo_id not in before_ids

        add_all = client.post("/api/v1/todos/my-day/add-all", headers=headers)
        assert add_all.status_code == 200
        assert add_all.json()["affected_count"] >= 2

        after_add_today = client.get("/api/v1/todos", headers=headers, params={"view": "TODAY"})
        assert after_add_today.status_code == 200
        added_ids = {item["id"] for item in after_add_today.json()}
        assert first_todo_id in added_ids
        assert second_todo_id in added_ids

        done_first = client.patch(
            f"/api/v1/todos/{first_todo_id}",
            headers=headers,
            json={"status": "DONE"},
        )
        assert done_first.status_code == 200
        assert done_first.json()["status"] == "DONE"

        today_after_done = client.get("/api/v1/todos", headers=headers, params={"view": "TODAY"})
        assert today_after_done.status_code == 200
        today_after_done_ids = {item["id"] for item in today_after_done.json()}
        assert first_todo_id not in today_after_done_ids
        assert second_todo_id in today_after_done_ids

        clear_today = client.post("/api/v1/todos/my-day/clear", headers=headers)
        assert clear_today.status_code == 200
        assert clear_today.json()["affected_count"] >= 1

        all_after_clear = client.get("/api/v1/todos", headers=headers, params={"view": "ALL", "include_done": True})
        assert all_after_clear.status_code == 200
        all_items = {item["id"]: item for item in all_after_clear.json()}
        # 未完成任务不会因“今日”清空而消失，仍留在全部任务里。
        assert second_todo_id in all_items
        assert all_items[second_todo_id]["status"] == "OPEN"


def test_periodic_billing_service_dates_auto_derive_due_date_and_month_range():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "自动推导账期客户",
                "contact_name": "推导负责人",
                "phone": "13800138131",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        monthly_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "amount_basis": "MONTHLY",
                "collection_start_date": "2026-04-15",
                "total_fee": 800,
                "monthly_fee": 800,
                "payment_method": "预收",
                "summary": "4月月费代账",
            },
        )
        assert monthly_resp.status_code == 201
        monthly_record = monthly_resp.json()
        assert monthly_record["collection_start_date"] == "2026-04-15"
        assert monthly_record["due_month"] == "2027-03-31"
        assert monthly_record["period_start_month"] == "2026-04"
        assert monthly_record["period_end_month"] == "2027-03"

        yearly_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "amount_basis": "YEARLY",
                "collection_start_date": "2026-04-15",
                "total_fee": 9600,
                "monthly_fee": 800,
                "payment_method": "预收",
                "summary": "年度代账",
            },
        )
        assert yearly_resp.status_code == 201
        yearly_record = yearly_resp.json()
        assert yearly_record["collection_start_date"] == "2026-04-15"
        assert yearly_record["due_month"] == "2027-03-31"
        assert yearly_record["period_start_month"] == "2026-04"
        assert yearly_record["period_end_month"] == "2027-03"



def test_one_time_billing_forces_one_time_basis_and_same_day_due_date():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "单次项目客户",
                "contact_name": "单次负责人",
                "phone": "13800138132",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_category": "注册",
                "charge_mode": "ONE_TIME",
                "amount_basis": "YEARLY",
                "collection_start_date": "2026-06-08",
                "total_fee": 3200,
                "payment_method": "预收",
                "summary": "股权变更一次性服务",
            },
        )
        assert record_resp.status_code == 201
        record = record_resp.json()
        assert record["amount_basis"] == "ONE_TIME"
        assert record["collection_start_date"] == "2026-06-08"
        assert record["due_month"] == "2026-06-08"
        assert record["period_start_month"] == ""
        assert record["period_end_month"] == ""



def test_periodic_billing_rejects_due_date_before_service_start_date():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "错误日期客户",
                "contact_name": "日期校验",
                "phone": "13800138133",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        create_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_mode": "PERIODIC",
                "amount_basis": "MONTHLY",
                "collection_start_date": "2026-08-15",
                "due_month": "2026-08-01",
                "total_fee": 1200,
                "payment_method": "后收",
            },
        )
        assert create_resp.status_code == 400
        assert create_resp.json()["detail"] == "到期日期不能早于服务开始日期"


def test_customer_detail_timeline_merges_pre_and_post_sale_records():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "时间线测试客户A",
                "contact_name": "王一",
                "phone": "13800138180",
                "source": "老板转介绍",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        followup_resp = client.post(
            f"/api/v1/leads/{lead_id}/followups",
            headers=headers,
            json={
                "followup_at": "2026-03-01",
                "feedback": "已首次联络并报价",
                "next_reminder_at": "2026-03-03",
                "notes": "等待客户确认是否下单",
            },
        )
        assert followup_resp.status_code == 201
        assert followup_resp.json()["created_by_username"] == "boss"

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        record_resp = client.post(
            "/api/v1/billing-records",
            headers=headers,
            json={
                "customer_id": customer_id,
                "charge_category": "代账",
                "charge_mode": "PERIODIC",
                "amount_basis": "MONTHLY",
                "collection_start_date": "2026-03-05",
                "due_month": "2027-03-04",
                "total_fee": 3600,
                "monthly_fee": 300,
                "payment_method": "后收",
                "summary": "代账首单",
            },
        )
        assert record_resp.status_code == 201
        record_id = record_resp.json()["id"]

        activity_resp = client.post(
            f"/api/v1/billing-records/{record_id}/activities",
            headers=headers,
            json={
                "activity_type": "PAYMENT",
                "occurred_at": "2026-03-06",
                "amount": 1000,
                "payment_nature": "ONE_OFF",
                "receipt_account": "聚能",
                "content": "首笔到账",
                "note": "客户先付一部分",
            },
        )
        assert activity_resp.status_code == 201
        assert activity_resp.json()["actor_username"] == "boss"
        assert activity_resp.json()["receipt_account"] == "聚能"

        execution_resp = client.post(
            f"/api/v1/billing-records/{record_id}/execution-logs",
            headers=headers,
            json={
                "occurred_at": "2026-03-07",
                "progress_type": "UPDATE",
                "content": "已收到执照并开始办理",
                "next_action": "等待补充章程",
                "note": "先走预审",
            },
        )
        assert execution_resp.status_code == 201

        event_resp = client.post(
            f"/api/v1/customers/{customer_id}/timeline-events",
            headers=headers,
            json={
                "occurred_at": "2026-03-08",
                "event_type": "MEETING",
                "content": "和老板讨论缺发票处理方式",
                "note": "决定先补票再报税",
                "amount": None,
            },
        )
        assert event_resp.status_code == 201
        assert event_resp.json()["actor_username"] == "boss"

        detail_resp = client.get(f"/api/v1/customers/{customer_id}", headers=headers)
        assert detail_resp.status_code == 200
        detail = detail_resp.json()
        source_types = {item["source_type"] for item in detail["timeline"]}
        assert "LEAD_FOLLOWUP" in source_types
        assert "CONVERTED" in source_types
        assert "BILLING_RECORD" in source_types
        assert "BILLING_ACTIVITY" in source_types
        assert "EXECUTION_LOG" in source_types
        assert "CUSTOMER_EVENT" in source_types

        manual_event = next(item for item in detail["timeline"] if item["source_type"] == "CUSTOMER_EVENT")
        assert manual_event["actor_username"] == "boss"
        assert manual_event["content"] == "和老板讨论缺发票处理方式"

        payment_event = next(item for item in detail["timeline"] if item["source_type"] == "BILLING_ACTIVITY")
        assert payment_event["amount"] == 1000
        assert payment_event["actor_username"] == "boss"


def test_customer_detail_keeps_development_source_separate_from_customer_records():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        users_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert users_resp.status_code == 200
        accountant = users_resp.json()[0]

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "name": "时间线测试客户B",
                "contact_name": "李二",
                "phone": "13800138181",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        source_followup_resp = client.post(
            f"/api/v1/leads/{lead_id}/followups",
            headers=headers,
            json={
                "followup_at": "2026-03-10",
                "feedback": "开发来源页的联络记录",
                "notes": "仅属于成单前",
            },
        )
        assert source_followup_resp.status_code == 201

        convert_resp = client.post(
            f"/api/v1/leads/{lead_id}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer_id = convert_resp.json()["customer"]["id"]

        customer_event_resp = client.post(
            f"/api/v1/customers/{customer_id}/timeline-events",
            headers=headers,
            json={
                "occurred_at": "2026-03-11",
                "event_type": "DOCUMENT",
                "content": "收到客户执照并安排后续变更",
                "note": "",
                "amount": None,
            },
        )
        assert customer_event_resp.status_code == 201

        lead_detail_resp = client.get(f"/api/v1/leads/{lead_id}/followups", headers=headers)
        assert lead_detail_resp.status_code == 200
        lead_followups = lead_detail_resp.json()
        assert len(lead_followups) == 1
        assert lead_followups[0]["feedback"] == "开发来源页的联络记录"

        customer_detail_resp = client.get(f"/api/v1/customers/{customer_id}", headers=headers)
        assert customer_detail_resp.status_code == 200
        customer_detail = customer_detail_resp.json()
        assert len(customer_detail["followups"]) == 1
        assert customer_detail["followups"][0]["feedback"] == "开发来源页的联络记录"
        assert any(
            item["source_type"] == "CUSTOMER_EVENT" and item["content"] == "收到客户执照并安排后续变更"
            for item in customer_detail["timeline"]
        )


def test_conversion_lead_defaults_contact_start_date_and_grade_reminder():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        create_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "CONVERSION",
                "name": "默认联络日期客户",
                "contact_name": "张三",
                "phone": "13800138182",
                "grade": "意向中",
            },
        )
        assert create_resp.status_code == 201
        body = create_resp.json()
        assert body["contact_start_date"] == date.today().isoformat()
        assert body["reminder_value"] == "7天"
        assert body["next_reminder_at"] == (date.today() + timedelta(days=7)).isoformat()


def test_conversion_lead_allows_blank_company_and_phone_with_contact_name_fallback():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        create_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "CONVERSION",
                "name": "",
                "contact_name": "王老板",
                "phone": "",
                "grade": "意向中",
                "main_business": "代账和报税咨询",
                "intro": "老客户转介绍",
            },
        )
        assert create_resp.status_code == 201
        body = create_resp.json()
        assert body["name"] == "王老板"
        assert body["phone"] == ""
        assert body["contact_start_date"] == date.today().isoformat()


def test_lead_intro_options_return_distinct_recent_matches():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        for index, intro in enumerate(["老客户转介绍", "百度渠道", "老客户转介绍", "抖音渠道"], start=1):
            create_resp = client.post(
                "/api/v1/leads",
                headers=headers,
                json={
                    "template_type": "CONVERSION",
                    "name": f"介绍人联想客户{index}",
                    "contact_name": f"联系人{index}",
                    "phone": "",
                    "grade": "意向中",
                    "main_business": "代理记账",
                    "intro": intro,
                },
            )
            assert create_resp.status_code == 201

        suggest_resp = client.get(
            "/api/v1/leads/intro-options",
            headers=headers,
            params={"q": "渠道", "limit": 10},
        )
        assert suggest_resp.status_code == 200
        assert suggest_resp.json() == ["抖音渠道", "百度渠道"]

        all_resp = client.get(
            "/api/v1/leads/intro-options",
            headers=headers,
            params={"limit": 10},
        )
        assert all_resp.status_code == 200
        all_items = all_resp.json()
        assert "老客户转介绍" in all_items
        assert "百度渠道" in all_items
        assert all_items.count("老客户转介绍") == 1


def test_lead_source_options_return_distinct_recent_matches():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        for index, source in enumerate(["Sally直播", "转介绍", "Sally直播", "抖音投流"], start=1):
            create_resp = client.post(
                "/api/v1/leads",
                headers=headers,
                json={
                    "template_type": "CONVERSION",
                    "name": f"来源联想客户{index}",
                    "contact_name": f"来源联系人{index}",
                    "phone": "",
                    "grade": "意向中",
                    "main_business": "代理记账",
                    "source": source,
                },
            )
            assert create_resp.status_code == 201

        suggest_resp = client.get(
            "/api/v1/leads/source-options",
            headers=headers,
            params={"q": "抖", "limit": 10},
        )
        assert suggest_resp.status_code == 200
        assert suggest_resp.json() == ["抖音投流"]

        all_resp = client.get(
            "/api/v1/leads/source-options",
            headers=headers,
            params={"limit": 10},
        )
        assert all_resp.status_code == 200
        all_items = all_resp.json()
        assert "Sally直播" in all_items
        assert "转介绍" in all_items
        assert all_items.count("Sally直播") == 1


def test_followup_updates_grade_and_reminder_defaults():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        lead_resp = client.post(
            "/api/v1/leads",
            headers=headers,
            json={
                "template_type": "CONVERSION",
                "name": "跟进等级测试客户",
                "contact_name": "李四",
                "phone": "13800138183",
            },
        )
        assert lead_resp.status_code == 201
        lead_id = lead_resp.json()["id"]

        followup_resp = client.post(
            f"/api/v1/leads/{lead_id}/followups",
            headers=headers,
            json={
                "followup_at": "2026-03-23",
                "grade": "待下单",
                "feedback": "客户正在等老板确认",
                "notes": "三天后再催",
            },
        )
        assert followup_resp.status_code == 201

        lead_detail_resp = client.get(f"/api/v1/leads/{lead_id}", headers=headers)
        assert lead_detail_resp.status_code == 200
        lead = lead_detail_resp.json()
        assert lead["grade"] == "待下单"
        assert lead["reminder_value"] == "3天"
        assert lead["next_reminder_at"] == "2026-03-26"
        assert lead["status"] == "FOLLOWING"

        lost_followup_resp = client.post(
            f"/api/v1/leads/{lead_id}/followups",
            headers=headers,
            json={
                "followup_at": "2026-03-26",
                "grade": "放弃",
                "feedback": "客户明确放弃",
                "notes": "不再跟进",
            },
        )
        assert lost_followup_resp.status_code == 201

        lost_lead_resp = client.get(f"/api/v1/leads/{lead_id}", headers=headers)
        assert lost_lead_resp.status_code == 200
        lost_lead = lost_lead_resp.json()
        assert lost_lead["grade"] == "放弃"
        assert lost_lead["reminder_value"] == "不跟进"
        assert lost_lead["next_reminder_at"] is None
        assert lost_lead["status"] == "LOST"


def test_customer_import_template_and_export_workbook():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        template_resp = client.get("/api/v1/customers/import-template", headers=headers)
        assert template_resp.status_code == 200
        template_wb = load_workbook(BytesIO(template_resp.content))
        assert template_wb.sheetnames == ["客户导入模板", "填写说明"]
        template_headers = [cell.value for cell in template_wb["客户导入模板"][1]]
        assert template_headers[:5] == ["客户ID", "客户编号", "公司名称", "联系人", "联系电话"]
        assert "会计账号" in template_headers

        export_resp = client.get("/api/v1/customers/export", headers=headers)
        assert export_resp.status_code == 200
        export_wb = load_workbook(BytesIO(export_resp.content))
        assert export_wb.active.title == "客户列表导出"
        export_headers = [cell.value for cell in export_wb.active[1]]
        assert export_headers == template_headers
        assert export_wb.active.max_row >= 2


def test_customer_import_creates_and_updates_customer():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        owner_headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        create_file = build_customer_import_bytes(
            [
                {
                    "name": "客户导入测试-A",
                    "contact_name": "导入联系人",
                    "phone": "13950004101",
                    "status": "ACTIVE",
                    "accountant_username": "accountant",
                    "grade": "A",
                    "region": "青岛",
                    "service_start_text": "2026-04-07",
                    "main_business": "代账服务",
                    "source": "Sally直播",
                    "intro": "麦总",
                    "fee_standard": "3600/年",
                    "first_billing_period": "2026-04",
                    "reminder_value": "7天",
                    "notes": "首次整表导入测试",
                }
            ]
        )

        import_resp = client.post(
            "/api/v1/customers/import",
            headers=owner_headers,
            files={
                "file": (
                    "customer-import.xlsx",
                    create_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert import_resp.status_code == 200
        import_result = import_resp.json()
        assert import_result["created_count"] == 1
        assert import_result["error_count"] == 0

        list_resp = client.get("/api/v1/customers", headers=owner_headers, params={"keyword": "客户导入测试-A"})
        assert list_resp.status_code == 200
        created_customer = next(item for item in list_resp.json() if item["name"] == "客户导入测试-A")
        assert created_customer["customer_code"].endswith("S")

        update_file = build_customer_import_bytes(
            [
                {
                    "customer_id": str(created_customer["id"]),
                    "customer_code": created_customer["customer_code"],
                    "name": "客户导入测试-A",
                    "contact_name": "导入联系人",
                    "phone": "13950004199",
                    "accountant_username": "accountant",
                    "fee_standard": "4200/年",
                    "notes": "导入更新成功",
                }
            ]
        )

        update_resp = client.post(
            "/api/v1/customers/import",
            headers=owner_headers,
            files={
                "file": (
                    "customer-import-update.xlsx",
                    update_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert update_resp.status_code == 200
        update_result = update_resp.json()
        assert update_result["updated_count"] == 1
        assert update_result["error_count"] == 0

        detail_resp = client.get(f"/api/v1/customers/{created_customer['id']}", headers=owner_headers)
        assert detail_resp.status_code == 200
        detail = detail_resp.json()
        assert detail["phone"] == "13950004199"
        assert detail["lead"]["fee_standard"] == "4200/年"
        assert detail["lead"]["notes"] == "导入更新成功"


def test_lead_list_supports_sorting_and_legacy_accountant_convert():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        seeds = [
            ("排序线索-C", "13800130031"),
            ("排序线索-A", "13800130032"),
            ("排序线索-B", "13800130033"),
        ]
        for name, phone in seeds:
            create_resp = client.post(
                "/api/v1/leads",
                headers=headers,
                json={
                    "name": name,
                    "contact_name": f"{name}-联系人",
                    "phone": phone,
                    "main_business": "排序测试",
                    "source": "Sally直播",
                },
            )
            assert create_resp.status_code == 201

        default_resp = client.get("/api/v1/leads", headers=headers, params={"keyword": "排序线索-"})
        assert default_resp.status_code == 200
        default_ids = [item["id"] for item in default_resp.json()]
        assert default_ids == sorted(default_ids, reverse=True)

        name_sorted_resp = client.get(
            "/api/v1/leads",
            headers=headers,
            params={"keyword": "排序线索-", "sort_by": "name", "sort_order": "asc"},
        )
        assert name_sorted_resp.status_code == 200
        assert [item["name"] for item in name_sorted_resp.json()] == ["排序线索-A", "排序线索-B", "排序线索-C"]

        accountants_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert accountants_resp.status_code == 200
        accountant = accountants_resp.json()[0]
        target_lead = next(item for item in name_sorted_resp.json() if item["name"] == "排序线索-A")

        convert_resp = client.post(
            f"/api/v1/leads/{target_lead['id']}/convert",
            headers=headers,
            json={"accountant_id": accountant["id"]},
        )
        assert convert_resp.status_code == 200
        customer = convert_resp.json()["customer"]
        assert customer["responsible_user_id"] == accountant["id"]
        assert customer["assigned_accountant_id"] == accountant["id"]


def test_customer_list_supports_sorting():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        accountants_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert accountants_resp.status_code == 200
        accountant = accountants_resp.json()[0]

        seeds = [
            ("排序客户-C", "13800130041"),
            ("排序客户-A", "13800130042"),
            ("排序客户-B", "13800130043"),
        ]
        for name, phone in seeds:
            lead_resp = client.post(
                "/api/v1/leads",
                headers=headers,
                json={
                    "name": name,
                    "contact_name": f"{name}-联系人",
                    "phone": phone,
                    "main_business": "客户排序测试",
                    "source": "Sally直播",
                },
            )
            assert lead_resp.status_code == 201
            convert_resp = client.post(
                f"/api/v1/leads/{lead_resp.json()['id']}/convert",
                headers=headers,
                json={"accountant_id": accountant["id"]},
            )
            assert convert_resp.status_code == 200

        default_resp = client.get("/api/v1/customers", headers=headers, params={"keyword": "排序客户-"})
        assert default_resp.status_code == 200
        default_ids = [item["id"] for item in default_resp.json()]
        assert default_ids == sorted(default_ids, reverse=True)

        name_sorted_resp = client.get(
            "/api/v1/customers",
            headers=headers,
            params={"keyword": "排序客户-", "sort_by": "name", "sort_order": "asc"},
        )
        assert name_sorted_resp.status_code == 200
        assert [item["name"] for item in name_sorted_resp.json()] == ["排序客户-A", "排序客户-B", "排序客户-C"]


def test_billing_record_list_supports_sorting():
    with TestClient(app) as client:
        owner_login = client.post(
            "/api/v1/auth/login",
            json={"username": "boss", "password": DEMO_PASSWORD},
        )
        assert owner_login.status_code == 200
        headers = {"Authorization": f"Bearer {owner_login.json()['access_token']}"}

        accountants_resp = client.get("/api/v1/users", headers=headers, params={"role": "ACCOUNTANT"})
        assert accountants_resp.status_code == 200
        accountant = accountants_resp.json()[0]

        seeds = [
            ("排序收费-C", "13800130051", 4003, "2026-12-31"),
            ("排序收费-A", "13800130052", 4001, "2026-10-31"),
            ("排序收费-B", "13800130053", 4002, "2026-11-30"),
        ]
        customer_ids: dict[str, int] = {}
        for name, phone, _, _ in seeds:
            lead_resp = client.post(
                "/api/v1/leads",
                headers=headers,
                json={
                    "name": name,
                    "contact_name": f"{name}-联系人",
                    "phone": phone,
                    "main_business": "收费排序测试",
                    "source": "Sally直播",
                },
            )
            assert lead_resp.status_code == 201
            convert_resp = client.post(
                f"/api/v1/leads/{lead_resp.json()['id']}/convert",
                headers=headers,
                json={"accountant_id": accountant["id"]},
            )
            assert convert_resp.status_code == 200
            customer_ids[name] = convert_resp.json()["customer"]["id"]

        for name, _, serial_no, due_month in seeds:
            create_resp = client.post(
                "/api/v1/billing-records",
                headers=headers,
                json={
                    "customer_id": customer_ids[name],
                    "serial_no": serial_no,
                    "total_fee": 1000,
                    "monthly_fee": 100,
                    "billing_cycle_text": "收费排序测试",
                    "due_month": due_month,
                    "payment_method": "后收",
                },
            )
            assert create_resp.status_code == 201

        default_resp = client.get("/api/v1/billing-records", headers=headers, params={"keyword": "排序收费-"})
        assert default_resp.status_code == 200
        assert [item["serial_no"] for item in default_resp.json()] == [4003, 4002, 4001]

        customer_sorted_resp = client.get(
            "/api/v1/billing-records",
            headers=headers,
            params={"keyword": "排序收费-", "sort_by": "customer_name", "sort_order": "asc"},
        )
        assert customer_sorted_resp.status_code == 200
        assert [item["customer_name"] for item in customer_sorted_resp.json()] == ["排序收费-A", "排序收费-B", "排序收费-C"]

        due_sorted_resp = client.get(
            "/api/v1/billing-records",
            headers=headers,
            params={"keyword": "排序收费-", "sort_by": "due_month", "sort_order": "asc"},
        )
        assert due_sorted_resp.status_code == 200
        assert [item["due_month"] for item in due_sorted_resp.json()] == ["2026-10-31", "2026-11-30", "2026-12-31"]


def test_auth_providers_and_local_login_transition_rules(monkeypatch):
    original = {
        "sso_enabled": settings.sso_enabled,
        "oidc_issuer": settings.oidc_issuer,
        "oidc_client_id": settings.oidc_client_id,
        "oidc_client_secret": settings.oidc_client_secret,
        "local_login_enabled": settings.local_login_enabled,
    }
    monkeypatch.setattr(settings, "sso_enabled", True)
    monkeypatch.setattr(settings, "oidc_issuer", "https://sso.example.com/realms/company")
    monkeypatch.setattr(settings, "oidc_client_id", "zhanghang-crm")
    monkeypatch.setattr(settings, "oidc_client_secret", "secret-value")
    monkeypatch.setattr(settings, "local_login_enabled", False)

    try:
        with TestClient(app) as client:
            providers_resp = client.get("/api/v1/auth/providers")
            assert providers_resp.status_code == 200
            providers = providers_resp.json()
            assert providers["sso"]["enabled"] is True
            assert providers["local"]["enabled"] is True
            assert providers["local"]["admin_only"] is True

            restricted_login = client.post(
                "/api/v1/auth/login",
                json={"username": "accountant", "password": DEMO_PASSWORD},
            )
            assert restricted_login.status_code == 403
            assert "企业单点登录" in restricted_login.json()["detail"]

            owner_login = client.post(
                "/api/v1/auth/login",
                json={"username": "boss", "password": DEMO_PASSWORD},
            )
            assert owner_login.status_code == 200
    finally:
        for key, value in original.items():
            setattr(settings, key, value)


def test_auth_me_supports_uid_claim_for_transition_tokens():
    with SessionLocal() as db:
        boss = db.execute(select(User).where(User.username == "boss")).scalar_one()
        token = create_access_token(subject="legacy-mismatch", role=boss.role, user_id=boss.id)

    with TestClient(app) as client:
        me_resp = client.get("/api/v1/auth/me", headers={"Authorization": f"Bearer {token}"})
        assert me_resp.status_code == 200
        assert me_resp.json()["username"] == "boss"


def test_sso_callback_exchange_can_issue_crm_token(monkeypatch):
    original = {
        "sso_enabled": settings.sso_enabled,
        "oidc_issuer": settings.oidc_issuer,
        "oidc_client_id": settings.oidc_client_id,
        "oidc_client_secret": settings.oidc_client_secret,
        "app_public_base_url": settings.app_public_base_url,
    }
    monkeypatch.setattr(settings, "sso_enabled", True)
    monkeypatch.setattr(settings, "oidc_issuer", "https://sso.example.com/realms/company")
    monkeypatch.setattr(settings, "oidc_client_id", "zhanghang-crm")
    monkeypatch.setattr(settings, "oidc_client_secret", "secret-value")
    monkeypatch.setattr(settings, "app_public_base_url", "https://ivanshang.com:26888")

    monkeypatch.setattr(auth_api, "sso_is_enabled", lambda: True)
    monkeypatch.setattr(
        auth_api,
        "exchange_code_for_tokens",
        lambda code: {"id_token": "fake-id-token", "access_token": "fake-access-token"},
    )
    verify_inputs = {}
    monkeypatch.setattr(
        auth_api,
        "verify_id_token",
        lambda token, nonce, access_token="": verify_inputs.update(
            {
                "token": token,
                "nonce": nonce,
                "access_token": access_token,
            }
        )
        or {
            "iss": "https://sso.example.com/realms/company",
            "sub": "keycloak-user-1",
            "preferred_username": "boss",
            "email": "boss@example.com",
        },
    )

    def fake_resolve_or_create_local_user(_db, _claims):
        return SimpleNamespace(id=1, username="boss", last_login_at=None), "BOUND"

    monkeypatch.setattr(auth_api, "resolve_or_create_local_user", fake_resolve_or_create_local_user)
    monkeypatch.setattr(
        auth_api,
        "get_valid_state_ticket",
        lambda _db, _state: SimpleNamespace(nonce="demo-nonce", status="PENDING", consumed_at=None),
    )
    monkeypatch.setattr(
        auth_api,
        "create_exchange_ticket",
        lambda _db, **_kwargs: SimpleNamespace(ticket="demo-exchange-ticket"),
    )
    monkeypatch.setattr(auth_api, "write_operation_log", lambda *args, **kwargs: None)

    try:
        with TestClient(app) as client:
            callback_resp = client.get(
                "/api/v1/auth/sso/callback",
                params={"state": "demo-state", "code": "demo-code"},
                follow_redirects=False,
            )
            assert callback_resp.status_code == 302
            location = callback_resp.headers["location"]
            assert "/login/sso?ticket=demo-exchange-ticket" in location
            assert verify_inputs["access_token"] == "fake-access-token"
    finally:
        for key, value in original.items():
            setattr(settings, key, value)


def test_admin_can_manage_sso_bindings_and_conflicts():
    manual_subject = f"manual-binding-{uuid4()}"
    conflict_subject = f"conflict-binding-{uuid4()}"
    with TestClient(app) as client:
        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        assert admin_login.status_code == 200
        headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}

        with SessionLocal() as db:
            boss = db.execute(select(User).where(User.username == "boss")).scalar_one()
            manager = db.execute(select(User).where(User.username == "manager")).scalar_one()
            existing_conflict = db.execute(
                select(SsoBindingConflict).where(
                    SsoBindingConflict.provider == "keycloak",
                    SsoBindingConflict.subject == conflict_subject,
                )
            ).scalar_one_or_none()
            if existing_conflict is None:
                existing_conflict = SsoBindingConflict(
                    provider="keycloak",
                    issuer="https://sso.example.com/realms/company",
                    subject=conflict_subject,
                    preferred_username="crm-test-conflict",
                    email="crm-test-conflict@example.com",
                    display_name="CRM 测试冲突",
                    raw_claims_json="{}",
                    reason="邮箱匹配到多个本地账号",
                    status="PENDING",
                    candidate_user_ids_json=f"[{manager.id}]",
                    first_seen_at=datetime.utcnow(),
                    last_seen_at=datetime.utcnow(),
                )
                db.add(existing_conflict)
                db.commit()
                db.refresh(existing_conflict)
            conflict_id = existing_conflict.id
            boss_id = boss.id
            manager_id = manager.id

        unbound_resp = client.get("/api/v1/admin/sso/unbound-users", headers=headers)
        assert unbound_resp.status_code == 200
        assert any(item["username"] == "boss" for item in unbound_resp.json())

        manual_bind_resp = client.post(
            "/api/v1/admin/sso/bindings/manual",
            headers=headers,
            json={
                "user_id": boss_id,
                "issuer": "https://sso.example.com/realms/company",
                "subject": manual_subject,
                "preferred_username": "crm-test",
                "email": f"crm-test-{uuid4()}@ivanshang.com",
                "display_name": "CRM 测试账号",
            },
        )
        assert manual_bind_resp.status_code == 201
        binding_id = manual_bind_resp.json()["id"]

        bindings_resp = client.get("/api/v1/admin/sso/bindings", headers=headers)
        assert bindings_resp.status_code == 200
        assert any(item["id"] == binding_id for item in bindings_resp.json())

        resolve_resp = client.post(
            f"/api/v1/admin/sso/conflicts/{conflict_id}/resolve",
            headers=headers,
            json={"user_id": manager_id},
        )
        assert resolve_resp.status_code == 200
        assert resolve_resp.json()["user_id"] == manager_id

        delete_resp = client.delete(f"/api/v1/admin/sso/bindings/{binding_id}", headers=headers)
        assert delete_resp.status_code == 204


def test_sso_exchange_ticket_is_one_time_use():
    with TestClient(app) as client:
        with SessionLocal() as db:
            ticket = auth_api.create_exchange_ticket(
                db,
                status="ERROR",
                error_code="DEMO",
                error_message="测试票据",
            )

        first_resp = client.post("/api/v1/auth/sso/exchange", json={"ticket": ticket.ticket})
        assert first_resp.status_code == 200
        assert first_resp.json()["status"] == "ERROR"

        second_resp = client.post("/api/v1/auth/sso/exchange", json={"ticket": ticket.ticket})
        assert second_resp.status_code == 400
        assert "票据已失效" in second_resp.json()["detail"]


def test_sso_exchange_pending_binding_uses_full_user_message():
    with TestClient(app) as client:
        with SessionLocal() as db:
            ticket = auth_api.create_exchange_ticket(
                db,
                status="CONFLICT",
                conflict_id=123,
            )

        exchange_resp = client.post("/api/v1/auth/sso/exchange", json={"ticket": ticket.ticket})
        assert exchange_resp.status_code == 200
        payload = exchange_resp.json()
        assert payload["status"] == "PENDING_BINDING"
        assert payload["conflict_id"] == 123
        assert payload["message"] == "当前企业账号需要管理员在后台确认绑定后才能进入 CRM。"


def test_verify_id_token_passes_access_token_for_at_hash(monkeypatch):
    original = {
        "sso_enabled": settings.sso_enabled,
        "oidc_issuer": settings.oidc_issuer,
        "oidc_client_id": settings.oidc_client_id,
        "oidc_client_secret": settings.oidc_client_secret,
    }
    monkeypatch.setattr(settings, "sso_enabled", True)
    monkeypatch.setattr(settings, "oidc_issuer", "https://sso.example.com/realms/company")
    monkeypatch.setattr(settings, "oidc_client_id", "zhanghang-crm")
    monkeypatch.setattr(settings, "oidc_client_secret", "secret-value")

    decode_calls = {}
    monkeypatch.setattr(sso_service.jwt, "get_unverified_header", lambda token: {"kid": "demo-kid", "alg": "RS256"})
    monkeypatch.setattr(sso_service, "_load_jwks", lambda: {"keys": [{"kid": "demo-kid", "kty": "RSA"}]})

    def fake_decode(token, key, algorithms=None, options=None, audience=None, issuer=None, subject=None, access_token=None):
        decode_calls["token"] = token
        decode_calls["audience"] = audience
        decode_calls["issuer"] = issuer
        decode_calls["access_token"] = access_token
        return {"nonce": "demo-nonce", "sub": "subject-1"}

    monkeypatch.setattr(sso_service.jwt, "decode", fake_decode)

    try:
        claims = sso_service.verify_id_token("id-token", nonce="demo-nonce", access_token="access-token")
        assert claims["sub"] == "subject-1"
        assert decode_calls["access_token"] == "access-token"
        assert decode_calls["audience"] == "zhanghang-crm"
        assert decode_calls["issuer"] == "https://sso.example.com/realms/company"
    finally:
        for key, value in original.items():
            setattr(settings, key, value)


def test_manual_sso_binding_resolves_matching_pending_conflict():
    subject = f"manual-resolve-{uuid4()}"
    with TestClient(app) as client:
        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        assert admin_login.status_code == 200
        headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}

        with SessionLocal() as db:
            manager = db.execute(select(User).where(User.username == "manager")).scalar_one()
            conflict = SsoBindingConflict(
                provider="keycloak",
                issuer="https://sso.example.com/realms/company",
                subject=subject,
                preferred_username="resolve-by-manual",
                email="resolve-by-manual@example.com",
                display_name="手动绑定自动消冲突",
                raw_claims_json="{}",
                reason="邮箱匹配到多个本地账号",
                status="PENDING",
                candidate_user_ids_json=f"[{manager.id}]",
                first_seen_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow(),
            )
            db.add(conflict)
            db.commit()
            db.refresh(conflict)
            conflict_id = conflict.id
            manager_id = manager.id

        manual_bind_resp = client.post(
            "/api/v1/admin/sso/bindings/manual",
            headers=headers,
            json={
                "user_id": manager_id,
                "issuer": "https://sso.example.com/realms/company",
                "subject": subject,
                "preferred_username": "resolve-by-manual",
                "email": "resolve-by-manual@example.com",
                "display_name": "手动绑定自动消冲突",
            },
        )
        assert manual_bind_resp.status_code == 201

        conflicts_resp = client.get("/api/v1/admin/sso/conflicts", headers=headers, params={"status_filter": "ALL"})
        assert conflicts_resp.status_code == 200
        matched = next(item for item in conflicts_resp.json() if item["id"] == conflict_id)
        assert matched["status"] == "RESOLVED"
        assert matched["resolved_user_id"] == manager_id


def test_sso_auto_bind_by_email_and_auto_create_projection():
    with TestClient(app):
        with SessionLocal() as db:
            accountant = db.execute(select(User).where(User.username == "accountant")).scalar_one()
            accountant.email = "accountant-sso@example.com"
            db.commit()

            claims_existing = {
                "iss": "https://sso.example.com/realms/company",
                "sub": f"auto-bind-email-{uuid4()}",
                "preferred_username": "some-other-login",
                "email": "accountant-sso@example.com",
                "email_verified": True,
                "name": "会计自动绑定",
            }
            user_existing, outcome_existing = sso_service.resolve_or_create_local_user(db, claims_existing)
            assert outcome_existing == "AUTO_BOUND_EMAIL"
            assert user_existing.id == accountant.id
            assert user_existing.external_managed is True
            assert any(identity.subject == claims_existing["sub"] for identity in user_existing.identities)

            claims_new = {
                "iss": "https://sso.example.com/realms/company",
                "sub": f"auto-create-{uuid4()}",
                "preferred_username": f"sso-new-{uuid4().hex[:8]}",
                "email": f"sso-new-{uuid4().hex[:8]}@example.com",
                "email_verified": True,
                "name": "SSO 新用户",
                "groups": ["crm-manager"],
            }
            user_new, outcome_new = sso_service.resolve_or_create_local_user(db, claims_new)
            assert outcome_new == "AUTO_CREATED"
            assert user_new.auth_source == "SSO"
            assert user_new.external_managed is True
            assert user_new.role == "MANAGER"
            assert user_new.email == claims_new["email"]


def test_sso_auto_bind_by_username_and_conflict_goes_pending():
    with TestClient(app):
        with SessionLocal() as db:
            claims_username = {
                "iss": "https://sso.example.com/realms/company",
                "sub": f"auto-bind-username-{uuid4()}",
                "preferred_username": "manager",
                "email": "",
                "name": "经理自动绑定",
            }
            user, outcome = sso_service.resolve_or_create_local_user(db, claims_username)
            assert outcome == "AUTO_BOUND_USERNAME"
            assert user.username == "manager"

            conflict_email = f"shared-{uuid4().hex[:8]}@example.com"
            boss = db.execute(select(User).where(User.username == "boss")).scalar_one()
            manager = db.execute(select(User).where(User.username == "manager")).scalar_one()
            boss.email = conflict_email
            manager.email = conflict_email
            db.commit()

            claims_conflict = {
                "iss": "https://sso.example.com/realms/company",
                "sub": f"conflict-{uuid4()}",
                "preferred_username": "conflict-user",
                "email": conflict_email,
                "name": "冲突用户",
            }
            try:
                sso_service.resolve_or_create_local_user(db, claims_conflict)
                assert False, "expected SsoConflictError"
            except sso_service.SsoConflictError as exc:
                conflict = exc.conflict
                assert conflict.status == "PENDING"
                assert conflict.reason == "邮箱匹配到多个本地账号"


def test_sso_binding_can_unbind_and_rebind_to_another_local_user():
    with TestClient(app) as client:
        admin_login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": DEMO_PASSWORD},
        )
        assert admin_login.status_code == 200
        headers = {"Authorization": f"Bearer {admin_login.json()['access_token']}"}

        subject = f"rebind-{uuid4()}"
        issuer = "https://sso.example.com/realms/company"

        with SessionLocal() as db:
            boss = db.execute(select(User).where(User.username == "boss")).scalar_one()
            manager = db.execute(select(User).where(User.username == "manager")).scalar_one()
            boss_id = boss.id
            manager_id = manager.id

        bind_first = client.post(
            "/api/v1/admin/sso/bindings/manual",
            headers=headers,
            json={
                "user_id": boss_id,
                "issuer": issuer,
                "subject": subject,
                "preferred_username": "rebind-user",
                "email": f"rebind-{uuid4().hex[:8]}@example.com",
                "display_name": "换绑测试用户",
            },
        )
        assert bind_first.status_code == 201
        binding_id = bind_first.json()["id"]
        assert bind_first.json()["user_id"] == boss_id

        unbind_resp = client.delete(f"/api/v1/admin/sso/bindings/{binding_id}", headers=headers)
        assert unbind_resp.status_code == 204

        bind_second = client.post(
            "/api/v1/admin/sso/bindings/manual",
            headers=headers,
            json={
                "user_id": manager_id,
                "issuer": issuer,
                "subject": subject,
                "preferred_username": "rebind-user",
                "email": f"rebind-second-{uuid4().hex[:8]}@example.com",
                "display_name": "换绑测试用户-新",
            },
        )
        assert bind_second.status_code == 201
        assert bind_second.json()["user_id"] == manager_id


def test_sso_exchange_fails_for_inactive_bound_user(monkeypatch):
    original = {
        "sso_enabled": settings.sso_enabled,
        "oidc_issuer": settings.oidc_issuer,
        "oidc_client_id": settings.oidc_client_id,
        "oidc_client_secret": settings.oidc_client_secret,
        "app_public_base_url": settings.app_public_base_url,
    }
    monkeypatch.setattr(settings, "sso_enabled", True)
    monkeypatch.setattr(settings, "oidc_issuer", "https://sso.example.com/realms/company")
    monkeypatch.setattr(settings, "oidc_client_id", "zhanghang-crm")
    monkeypatch.setattr(settings, "oidc_client_secret", "secret-value")
    monkeypatch.setattr(settings, "app_public_base_url", "https://ivanshang.com:26888")

    monkeypatch.setattr(auth_api, "sso_is_enabled", lambda: True)
    monkeypatch.setattr(
        auth_api,
        "build_sso_login_url",
        lambda ticket: f"https://sso.example.com/auth?state={ticket.ticket}&nonce={ticket.nonce}",
    )
    monkeypatch.setattr(
        auth_api,
        "exchange_code_for_tokens",
        lambda code: {"id_token": "fake-id-token", "access_token": "fake-access-token"},
    )
    monkeypatch.setattr(
        auth_api,
        "verify_id_token",
        lambda token, nonce, access_token="": {
            "iss": "https://sso.example.com/realms/company",
            "sub": "inactive-user-subject",
            "preferred_username": "accountant",
            "email": "inactive-bound@example.com",
        },
    )

    def fake_resolve_inactive(db, claims):
        user = db.execute(select(User).where(User.username == "accountant")).scalar_one()
        user.is_active = False
        db.commit()
        return user, "BOUND"

    monkeypatch.setattr(auth_api, "resolve_or_create_local_user", fake_resolve_inactive)

    try:
        with TestClient(app) as client:
            login_resp = client.get("/api/v1/auth/sso/login", follow_redirects=False)
            assert login_resp.status_code == 302
            state = parse_qs(urlparse(login_resp.headers["location"]).query)["state"][0]

            callback_resp = client.get(
                "/api/v1/auth/sso/callback",
                params={"state": state, "code": "demo-code"},
                follow_redirects=False,
            )
            assert callback_resp.status_code == 302
            exchange_ticket = callback_resp.headers["location"].split("ticket=", 1)[1]

            exchange_resp = client.post("/api/v1/auth/sso/exchange", json={"ticket": exchange_ticket})
            assert exchange_resp.status_code == 401
    finally:
        with SessionLocal() as db:
            accountant = db.execute(select(User).where(User.username == "accountant")).scalar_one()
            accountant.is_active = True
            db.commit()
        for key, value in original.items():
            setattr(settings, key, value)
